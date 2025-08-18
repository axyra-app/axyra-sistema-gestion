# backend/utils.py
"""
Funciones auxiliares para el sistema de nómina y cuadre de caja
"""

import sqlite3
import os
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle, Paragraph, Spacer, SimpleDocTemplate
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from datetime import datetime
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from google.auth.transport.requests import Request
from googleapiclient.http import MediaFileUpload
import json
import shutil
from typing import List, Optional
from pydantic import BaseModel
import pytz

# === Rutas ===
ROOT_DIR = os.path.dirname(os.path.dirname(__file__))
DB_PATH = os.path.join(ROOT_DIR, "data", "nomina.db")
PDF_DIR = os.path.join(os.path.dirname(__file__), "pdfs")
EXCEL_DIR = os.path.join(ROOT_DIR, "plantillas")
os.makedirs(PDF_DIR, exist_ok=True)
os.makedirs(EXCEL_DIR, exist_ok=True)

# Importar get_db desde database para evitar conflictos
from database import get_db

# === Configuración por defecto ===
def cargar_configuracion():
    """Cargar configuración desde archivo JSON."""
    try:
        config_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "config.json")
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        return config
    except Exception as e:
        # Error al cargar configuración
        # Configuración por defecto en pesos colombianos
        return {
            "empresa": {
                "nombre": "VILLA VENECIA",
                "nit": "900.123.456-7",
                "direccion": "Calle 123 # 45-67, Bogotá D.C."
            },
            "salario_minimo": 1423500,
            "auxilio_transporte": 100000,
            "porcentajes": {
                "salud": 0.04,
                "pension": 0.04,
                "fondo_solidaridad": 0.0,
                "retencion_fuente": 0.0
            },
            "recargos": {
                "nocturno": 0.35,
                "dominical": 0.75,
                "nocturno_dominical": 1.10,
                "extra_diurna": 1.25,
                "extra_nocturna": 1.75,
                "diurna_dominical": 0.80,
                "nocturna_dominical": 1.10,
                "extra_diurna_dominical": 1.05,
                "extra_nocturna_dominical": 1.85
            },
            "moneda": "COP",
            "formato_moneda": "es-CO"
        }

# === Tipos de horas ===
TIPOS_HORAS = [
    ("ordinarias", 0.00),
    ("recargo_nocturno", 0.35),
    ("recargo_diurno_dominical", 0.75),
    ("recargo_nocturno_dominical", 1.10),
    ("hora_extra_diurna", 0.25),
    ("hora_extra_nocturna", 0.75),
    ("hora_diurna_dominical_o_festivo", 0.80),
    ("hora_extra_diurna_dominical_o_festivo", 1.05),
    ("hora_nocturna_dominical_o_festivo", 1.10),
    ("hora_extra_nocturna_dominical_o_festivo", 1.85)
]

# === Base de datos: Conexión ===
# La función get_db se importa desde database.py para evitar conflictos

def calcular_auxilio_transporte(salario, salario_devengado):
    """Calcula el auxilio de transporte según la fórmula: =SI(D8<=1500000*2; (200000/30)*(F8/(D8/30)); 0)"""
    salario_minimo_doble = 1500000 * 2  # 3,000,000
    if salario <= salario_minimo_doble:
        # (200000/30) * (salario_devengado / (salario/30))
        # Simplificado: (200000 * salario_devengado) / salario
        return (200000 * salario_devengado) / salario if salario > 0 else 0
    else:
        return 0

def es_empleado_fijo(tipo_empleado):
    """Determina si un empleado es fijo o temporal."""
    return tipo_empleado.upper() == 'FIJO'

def aplicar_deducciones_empleado(tipo_empleado, salario_devengado, config):
    """Aplica deducciones según el tipo de empleado."""
    if es_empleado_fijo(tipo_empleado):
        # Empleados fijos: Salud, Pensión
        salud = salario_devengado * config['porcentajes']['salud']
        pension = salario_devengado * config['porcentajes']['pension']
        return salud, pension
    else:
        # Empleados temporales: No tienen deducciones
        return 0, 0

def calcular_auxilio_transporte_empleado(tipo_empleado, salario, salario_devengado):
    """Calcula auxilio de transporte según el tipo de empleado."""
    if es_empleado_fijo(tipo_empleado):
        return calcular_auxilio_transporte(salario, salario_devengado)
    else:
        # Empleados temporales no reciben auxilio de transporte
        return 0

# === Obtener empleados ===
def obtener_empleados(conn=None):
    close_conn = False
    if conn is None:
        conn = get_db()
        close_conn = True
    cursor = conn.cursor()
    cursor.execute("SELECT id, nombre, cedula, tipo, salario FROM empleados ORDER BY tipo, nombre")
    empleados = cursor.fetchall()
    
    # Convertir a lista de diccionarios con formato correcto
    empleados_formateados = []
    for empleado in empleados:
        # Mapear el tipo de la base de datos al formato del frontend
        tipo_frontend = "Fijo" if empleado[3] == "FIJO" else "Por Horas"
        empleados_formateados.append({
            "id": empleado[0],
            "nombre": empleado[1],
            "cedula": empleado[2],
            "tipo_contrato": tipo_frontend,
            "salario": empleado[4]
        })
    
    if close_conn:
        conn.close()
    return empleados_formateados

# === Guardar empleado ===
def guardar_empleado(nombre, cedula, tipo_contrato, salario):
    conn = get_db()
    cursor = conn.cursor()
    
    # Mapear los valores del formulario a los valores de la base de datos
    tipo_db = "FIJO" if tipo_contrato == "Fijo" else "TEMPORAL"
    
    try:
        cursor.execute(
            "INSERT INTO empleados (nombre, cedula, tipo, salario) VALUES (?, ?, ?, ?)",
            (nombre, cedula, tipo_db, salario)
        )
        conn.commit()
    except sqlite3.IntegrityError:
        raise ValueError(f"Ya existe un empleado con cédula {cedula}")
    finally:
        conn.close()


# === Actualizar empleado ===
def actualizar_empleado(empleado_id, nombre, cedula, tipo_contrato, salario):
    conn = get_db()
    cursor = conn.cursor()
    
    # Mapear los valores del formulario a los valores de la base de datos
    tipo_db = "FIJO" if tipo_contrato == "Fijo" else "TEMPORAL"
    
    try:
        cursor.execute(
            "UPDATE empleados SET nombre = ?, cedula = ?, tipo = ?, salario = ? WHERE id = ?",
            (nombre, cedula, tipo_db, salario, empleado_id)
        )
        conn.commit()
        
        if cursor.rowcount == 0:
            raise ValueError(f"No se encontró el empleado con ID {empleado_id}")
    except sqlite3.IntegrityError:
        raise ValueError(f"Ya existe un empleado con cédula {cedula}")
    finally:
        conn.close()




# === Guardar horas trabajadas ===
def guardar_horas_trabajadas(
    empleado_id, quincena, horas_ordinarias, recargo_nocturno, recargo_diurno_dominical,
    recargo_nocturno_dominical, hora_extra_diurna, hora_extra_nocturna,
    hora_diurna_dominical_o_festivo, hora_extra_diurna_dominical_o_festivo,
    hora_nocturna_dominical_o_festivo, hora_extra_nocturna_dominical_o_festivo,
    motivo_deuda="", valor_deuda=0.0, descuento_inasistencia=0.0, 
    otros_descuentos="", valor_otros_descuentos=0.0
):
    conn = get_db()
    cursor = conn.cursor()
    
    # Usar INSERT OR REPLACE para la nueva estructura
    cursor.execute("""
        INSERT OR REPLACE INTO horas_trabajadas (
            empleado_id, quincena, horas_ordinarias, recargo_nocturno, 
            recargo_diurno_dominical, recargo_nocturno_dominical, hora_extra_diurna,
            hora_extra_nocturna, hora_diurna_dominical_o_festivo,
            hora_extra_diurna_dominical_o_festivo, hora_nocturna_dominical_o_festivo,
            hora_extra_nocturna_dominical_o_festivo, motivo_deuda, valor_deuda,
            descuento_inasistencia, otros_descuentos, valor_otros_descuentos
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        empleado_id, quincena, horas_ordinarias, recargo_nocturno,
        recargo_diurno_dominical, recargo_nocturno_dominical, hora_extra_diurna,
        hora_extra_nocturna, hora_diurna_dominical_o_festivo,
        hora_extra_diurna_dominical_o_festivo, hora_nocturna_dominical_o_festivo,
        hora_extra_nocturna_dominical_o_festivo, motivo_deuda, valor_deuda,
        descuento_inasistencia, otros_descuentos, valor_otros_descuentos
    ))
    
    conn.commit()
    conn.close()

# === Calcular nómina por empleado ===
def calcular_valores(tipo: str, salario_base: float, horas_dict: dict, config: dict, valor_deuda: float = 0):
    """Calcular valores de nómina para un empleado con validaciones."""
    # Usar el salario base del empleado dividido por 220
    valor_hora_base = salario_base / 220
    
    total = 0
    detalle = []
    total_horas = 0

    # Calcular total de horas trabajadas (excluyendo ordinarias)
    total_horas = sum(horas_dict.values()) - horas_dict.get("ordinarias", 0)

    # Horas ordinarias (solo las horas ordinarias, no el total)
    horas_ordinarias = horas_dict.get("ordinarias", 0)
    if horas_ordinarias and horas_ordinarias > 0:
        if tipo == "FIJO" and horas_ordinarias > 88:
            horas_ordinarias = 88  # Límite para empleados fijos
        
        # Para empleados fijos, el salario devengado debe ser salario mínimo / 2
        if tipo == "FIJO":
            salario_minimo = config.get("salario_minimo", 1423500)
            subtotal_ordinarias = salario_minimo / 2
            # Empleado fijo: salario_devengado = salario_minimo/2
        else:
            subtotal_ordinarias = round(horas_ordinarias * valor_hora_base, 2)
            # Empleado temporal: salario_devengado = horas * valor_hora_base
        
        total += subtotal_ordinarias
        detalle.append(("Salario Devengado", 0.00, round(valor_hora_base, 2), horas_ordinarias, 0, subtotal_ordinarias))
    else:
        # Agregar salario devengado con 0 horas
        if tipo == "FIJO":
            salario_minimo = config.get("salario_minimo", 1423500)
            subtotal_ordinarias = salario_minimo / 2
            # Empleado fijo sin horas: salario_devengado = salario_minimo/2
        else:
            subtotal_ordinarias = 0
            # Empleado temporal sin horas: salario_devengado = 0
        
        detalle.append(("Salario Devengado", 0.00, round(valor_hora_base, 2), 0, 0, subtotal_ordinarias))

    # Recargos - AGREGAR TODOS LOS CONCEPTOS, incluso con 0 horas
    recargos_config = {
        "recargo_nocturno": ("Recargo Nocturno", 0.35),
        "recargo_diurno_dominical": ("Recargo dominical laborado", 0.75),
        "recargo_nocturno_dominical": ("Recargo Nocturno dominical o festivo", 1.10),
        "hora_extra_diurna": ("Hora extra diurna", 0.25),
        "hora_extra_nocturna": ("Hora extra nocturna", 0.75),
        "hora_diurna_dominical_o_festivo": ("Hora diurna dominical o festivo", 0.80),
        "hora_extra_diurna_dominical_o_festivo": ("Hora extra diurna dominical o festivo", 1.10),
        "hora_nocturna_dominical_o_festivo": ("Hora nocturna Dominical o festiva", 1.05),
        "hora_extra_nocturna_dominical_o_festivo": ("Hora extra nocturna dominical o festivo", 1.85)
    }

    # AGREGAR TODOS LOS RECARGOS, incluso con 0 horas
    for concepto_key, (concepto_nombre, porcentaje) in recargos_config.items():
        horas = horas_dict.get(concepto_key, 0)
        
        # Ocultar "Hora extra diurna" solo si es un fijo y no hay horas
        if tipo == "FIJO" and concepto_key == "hora_extra_diurna" and horas == 0:
            continue
        
        # Validar límites de horas extras si hay horas
        if horas and horas > 0:
            if "extra" in concepto_key and tipo == "FIJO":
                if horas > 20:  # Máximo 20 horas extras por quincena
                    horas = 20
            
            valor_recargo = valor_hora_base * porcentaje
            valor_total_hora = valor_hora_base + valor_recargo
            subtotal = round(horas * valor_total_hora, 2)
            
            # Validar que el subtotal no sea excesivo
            if subtotal > salario_base * 0.5:  # Máximo 50% del salario base
                subtotal = round(salario_base * 0.5, 2)
            
            total += subtotal
        else:
            # Concepto con 0 horas - mostrar valores pero subtotal 0
            valor_recargo = valor_hora_base * porcentaje
            valor_total_hora = valor_hora_base + valor_recargo
            subtotal = 0
            horas = 0
        
        # Agregar a detalle con el orden correcto: (concepto, porcentaje, valor_hora_base, horas, valor_recargo, subtotal)
        detalle.append((concepto_nombre, porcentaje, round(valor_hora_base, 2), horas, round(valor_recargo, 2), subtotal))

    # Auxilio de transporte (solo fijos con salario bajo)
    auxilio = 0
    if tipo == "FIJO" and salario_base <= 2 * config["salario_minimo"] and horas_ordinarias > 0:
        auxilio = config["auxilio_transporte"]

    # Deducciones (solo fijos)
    salud = 0
    pension = 0
    if tipo == "FIJO":
        salud = round(total * config["porcentajes"]["salud"], 2)
        pension = round(total * config["porcentajes"]["pension"], 2)
    
    # Manejar valor_deuda de forma segura
    try:
        deuda = float(valor_deuda) if valor_deuda else 0
        # Limitar deuda al 30% del salario base
        if deuda > salario_base * 0.3:
            deuda = round(salario_base * 0.3, 2)
    except (ValueError, TypeError):
        deuda = 0
    
    # Calcular neto con validaciones
    neto = total + auxilio - salud - pension - deuda
    
    # Asegurar que el neto no sea negativo
    if neto < 0:
        neto = 0

    resumen = {
        "total": round(total, 2),
        "auxilio": round(auxilio, 2),
        "salud": round(salud, 2),
        "pension": round(pension, 2),
        "deuda": round(deuda, 2),
        "neto": round(neto, 2)
    }

    return detalle, resumen

# === Calcular nóminas para todos los empleados ===
def calcular_nominas(empleados, quincena):
    """Calcular nómina para una lista de empleados."""
    conn = get_db()
    cursor = conn.cursor()
    config = cargar_configuracion()
    nominas = []
    
    for emp in empleados:
        cursor.execute("""
            SELECT horas_ordinarias, recargo_nocturno, recargo_diurno_dominical,
                   recargo_nocturno_dominical, hora_extra_diurna, hora_extra_nocturna,
                   hora_diurna_dominical_o_festivo, hora_extra_diurna_dominical_o_festivo,
                   hora_nocturna_dominical_o_festivo, hora_extra_nocturna_dominical_o_festivo, motivo_deuda, valor_deuda
            FROM horas_trabajadas
            WHERE empleado_id = ? AND quincena = ?
        """, (emp["id"], quincena))
        
        horas_row = cursor.fetchone()
        if horas_row:
            horas_dict = {
                "ordinarias": horas_row[0] or 0,
                "recargo_nocturno": horas_row[1] or 0,
                "recargo_diurno_dominical": horas_row[2] or 0,
                "recargo_nocturno_dominical": horas_row[3] or 0,
                "hora_extra_diurna": horas_row[4] or 0,
                "hora_extra_nocturna": horas_row[5] or 0,
                "hora_diurna_dominical_o_festivo": horas_row[6] or 0,
                "hora_extra_diurna_dominical_o_festivo": horas_row[7] or 0,
                "hora_nocturna_dominical_o_festivo": horas_row[8] or 0,
                "hora_extra_nocturna_dominical_o_festivo": horas_row[9] or 0
            }
            valor_deuda = horas_row[10] or 0
        else:
            horas_dict = {th: 0 for th, _ in TIPOS_HORAS}
            valor_deuda = 0

        detalle, resumen = calcular_valores(emp["tipo"], emp["salario"], horas_dict, config, valor_deuda)
        nominas.append({
            "empleado": emp,
            "detalle": detalle,
            "resumen": resumen,
            "total_horas": sum(row[3] for row in detalle)  # Cambiado de row[4] a row[3] porque las horas están en el índice 3
        })
    
    conn.close()
    return nominas

def generar_comprobante_pdf(empleado_id, quincena):
    """
    Genera un PDF de comprobante de nómina usando la misma estructura del código de escritorio
    
    Args:
        empleado_id: ID del empleado en la base de datos
        quincena: Quincena de trabajo (formato: "15_MM_YYYY" o "31_MM_YYYY")
    """
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM empleados WHERE id = ?", (empleado_id,))
    emp = cursor.fetchone()
    if not emp:
        raise ValueError("Empleado no encontrado")

    cursor.execute("""
        SELECT horas_ordinarias, recargo_nocturno, recargo_diurno_dominical,
               recargo_nocturno_dominical, hora_extra_diurna, hora_extra_nocturna,
               hora_diurna_dominical_o_festivo, hora_extra_diurna_dominical_o_festivo,
               hora_nocturna_dominical_o_festivo, hora_extra_nocturna_dominical_o_festivo, motivo_deuda, valor_deuda
        FROM horas_trabajadas
        WHERE empleado_id = ? AND quincena = ?
    """, (empleado_id, quincena))
    
    horas_row = cursor.fetchone()
    conn.close()
    
    if not horas_row:
        raise ValueError("No hay datos de horas para generar el PDF")
    
    # Construir el diccionario de horas como lo espera tu función original
    horas_dict = {
        "ordinarias": horas_row[0] or 0,
        "recargo_nocturno": horas_row[1] or 0,
        "recargo_diurno_dominical": horas_row[2] or 0,
        "recargo_nocturno_dominical": horas_row[3] or 0,
        "hora_extra_diurna": horas_row[4] or 0,
        "hora_extra_nocturna": horas_row[5] or 0,
        "hora_diurna_dominical_o_festivo": horas_row[6] or 0,
        "hora_extra_diurna_dominical_o_festivo": horas_row[7] or 0,
        "hora_nocturna_dominical_o_festivo": horas_row[8] or 0,
        "hora_extra_nocturna_dominical_o_festivo": horas_row[9] or 0
    }
    
    # Calcular total de horas
    horas_dict["total_horas"] = sum(horas_dict.values())
    
    deuda_comentario = horas_row[10] or ""
    deuda_valor = horas_row[11] or 0
    
    # Llamar a la función principal de generación de PDF
    return generar_pdf(
        nombre=emp[1],
        cedula=emp[2], 
        tipo=emp[3],
        salario_base=emp[4],
        horas_dict=horas_dict,
        config=cargar_configuracion(),
        deuda_comentario=deuda_comentario,
        deuda_valor=deuda_valor
    )


def generar_pdf(nombre, cedula, tipo, salario_base, horas_dict, config, deuda_comentario="", deuda_valor=0):
    """
    Genera un PDF de comprobante de nómina
    
    Args:
        nombre: Nombre del empleado
        cedula: Cédula del empleado
        tipo: Tipo de contrato (FIJO o TEMPORAL)
        salario_base: Salario base del empleado
        horas_dict: Diccionario con las horas trabajadas
        config: Configuración del sistema
        deuda_comentario: Comentario sobre la deuda (opcional)
        deuda_valor: Valor de la deuda (opcional)
    """
    # Crear directorio si no existe
    if not os.path.exists(PDF_DIR):
        try:
            os.makedirs(PDF_DIR)
        except OSError as e:
            # Error creando directorio PDF
            raise OSError(f"No se pudo crear el directorio {PDF_DIR}: {str(e)}")
    
    # Generar nombre de archivo seguro
    nombre_seguro = "".join(c if c.isalnum() or c in " _-" else "_" 
                           for c in nombre.upper().replace(" ", "_"))
    fecha = datetime.now().strftime("%m-%d-%Y_%I-%M-%p").replace(":", "-")
    archivo = os.path.join(PDF_DIR, f"COMPROBANTE_{nombre_seguro}_{cedula}_{fecha}.pdf")
    
    try:
        # Crear documento PDF
        doc = SimpleDocTemplate(archivo, 
                              pagesize=letter, 
                              rightMargin=36, 
                              leftMargin=72, 
                              topMargin=72, 
                              bottomMargin=72)
        elements = []
        styles = getSampleStyleSheet()
        
        # Estilo personalizado
        custom_style = styles["Normal"].clone('CustomStyle')
        custom_style.fontName = "Helvetica"
        custom_style.fontSize = 10
        custom_style.leading = 12
        custom_style.textColor = "#000000"
        custom_style.alignment = 0
        
        # Encabezado
        elements.append(Paragraph(f"ORDEN DE TRABAJO N° {fecha.upper()}", styles["Heading1"]))
        elements.append(Spacer(1, 12))
        
        # Información de la empresa
        elements.append(Paragraph(f"EMPRESA: {config['empresa']['nombre'].upper()}", custom_style))
        elements.append(Paragraph(f"NIT: {config['empresa']['nit'].upper()}", custom_style))
        elements.append(Paragraph(f"DIRECCIÓN: {config['empresa']['direccion'].upper()}", custom_style))
        elements.append(Spacer(1, 12))
        
        # Información del empleado
        elements.append(Paragraph(f"PRESTADOR DEL SERVICIO: {nombre.upper()}", custom_style))
        elements.append(Paragraph(f"CÉDULA: {cedula.upper()}", custom_style))
        elements.append(Paragraph(f"TIPO DE CONTRATO: {tipo.upper()}", custom_style))
        elements.append(Paragraph(f"SALARIO BASE LIQUIDACIÓN: ${salario_base:,.0f}", custom_style))
        elements.append(Paragraph(f"TOTAL HORAS TRABAJADAS: {horas_dict['total_horas']:.1f}", custom_style))
        elements.append(Spacer(1, 12))
        
        # Tabla de conceptos
        table_data = [
            ["CONCEPTO", "VALOR\nHORA", "VALOR\nRECARGO", "VALOR HORA\nCON RECARGO", "HORAS", "SUBTOTAL"],
        ]
        
        # Calcular valores
        detalle, resumen = calcular_valores(tipo, salario_base, horas_dict, config)
        if not detalle:
            raise ValueError("No hay datos de horas para generar el PDF")
        
        # Agregar filas de conceptos
        for row in detalle:
            concept = row[0]
            
            # Ocultar "Horas extras diurnas" solo si es un fijo
            if tipo == "FIJO" and concept.strip().lower() == "horas extras diurnas":
                continue
            
            porcentaje = row[1]
            valor_hora_base = row[2]
            horas = row[3]
            valor_recargo = row[4]
            subtotal = row[5]
            
            table_data.append([
                concept.upper(),
                f"${valor_hora_base:,.0f}",
                f"${valor_recargo:,.0f}",
                f"${valor_hora_base + valor_recargo:,.0f}",
                f"{horas:.1f}",
                f"${subtotal:,.0f}"
            ])
        
        # Crear tabla
        row_heights = [20] + [20] * (len(table_data) - 1)
        table = Table(table_data, 
                     colWidths=[120, 80, 70, 90, 50, 80], 
                     rowHeights=row_heights)
        
        # Aplicar estilos a la tabla
        table.setStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('LEADING', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (5, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
            ('ALIGN', (1, 0), (3, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ])
        
        elements.append(table)
        elements.append(Spacer(1, 12))
        
        # Resumen de valores
        total = resumen["total"]
        auxilio = resumen["auxilio"]
        salud = resumen["salud"]
        pension = resumen["pension"]
        deuda = deuda_valor if deuda_valor else 0
        neto = resumen["neto"]
        total_con_auxilio = total + auxilio
        
        if auxilio > 0:
            elements.append(Paragraph(f"AUXILIO DE TRANSPORTE: ${auxilio:,.0f}", custom_style))
        
        if tipo == "FIJO":
            elements.append(Paragraph(f"APORTE SALUD: (${salud:,.0f})", custom_style))
            elements.append(Paragraph(f"APORTE PENSIÓN: (${pension:,.0f})", custom_style))
        
        if deuda_valor > 0:
            elements.append(Paragraph(f"MOTIVO DE DEUDA: {deuda_comentario.upper()}", custom_style))
            elements.append(Paragraph(f"DEUDA: (${deuda_valor:,.0f})", custom_style))
        
        elements.append(Paragraph(f"TOTAL (CON AUXILIO): ${total_con_auxilio:,.0f}", custom_style))
        neto_ajustado = max(0, neto - deuda)
        elements.append(Paragraph(f"TOTAL NETO A PAGAR: ${neto_ajustado:,.0f}", custom_style))
        elements.append(Spacer(1, 24))
        
        # Firmas
        elements.append(Paragraph("FIRMA DEL TRABAJADOR: __________________________", custom_style))
        elements.append(Paragraph(f"CÉDULA: __________________________", custom_style))
        elements.append(Spacer(1, 12))
        
        # Construir PDF
        doc.build(elements)
        # PDF generado exitosamente
        return archivo
        
    except Exception as e:
        # Error generando PDF
        raise Exception(f"No se pudo generar el PDF: {str(e)}")

def generar_pdf_fijo(doc, emp, quincena, detalle, resumen, config, horas_dict, valor_deuda):
    """Generar PDF para empleados fijos con estructura clara y organizada."""
    width, height = letter
    
    # === ENCABEZADO PRINCIPAL ===
    doc.setFont("Helvetica-Bold", 18)
    doc.drawString(200, height - 50, "ORDEN DE TRABAJO")
    
    # === INFORMACIÓN DE LA EMPRESA ===
    doc.setFont("Helvetica-Bold", 12)
    doc.drawString(100, height - 80, f"EMPRESA: {config['empresa']['nombre']}")
    doc.drawString(100, height - 100, f"NIT: {config['empresa']['nit']}")
    doc.drawString(100, height - 120, f"DIRECCIÓN: {config['empresa']['direccion']}")
    
    # === INFORMACIÓN DEL EMPLEADO ===
    doc.setFont("Helvetica-Bold", 12)
    doc.drawString(100, height - 150, f"EMPLEADO: {emp[1].upper()}")
    doc.drawString(100, height - 170, f"CÉDULA: {emp[2]}")
    
    # === PERIODO DE TRABAJO ===
    # La fecha seleccionada solo afecta fecha inicio y fecha fin
    # No es un comprobante por día, es por quincena
    try:
        if "_" in quincena:
            partes = quincena.split("_")
            if len(partes) == 3:
                dia, mes, año = partes
                # Para quincenas: del 1 al 15 o del 16 al 31
                if int(dia) <= 15:
                    fecha_inicio = f"01/{mes}/{año}"
                    fecha_fin = f"15/{mes}/{año}"
                else:
                    fecha_inicio = f"16/{mes}/{año}"
                    fecha_fin = f"31/{mes}/{año}"
            else:
                fecha_inicio = fecha_fin = quincena
        else:
            fecha_inicio = fecha_fin = quincena
    except:
        fecha_inicio = fecha_fin = quincena
    
    doc.drawString(100, height - 190, f"FECHA DE INICIO: {fecha_inicio}")
    doc.drawString(100, height - 210, f"FECHA DE TERMINACIÓN: {fecha_fin}")
    
    # === HORAS Y SALARIO ===
    total_horas = sum(horas_dict.values())
    doc.drawString(100, height - 230, f"Horas prestación de servicio: {total_horas}")
    doc.drawString(100, height - 250, f"SALARIO BASE LIQUIDACIÓN: $ {emp[4]:,.0f}")
    
    # === TABLA PRINCIPAL DE CONCEPTOS ===
    y = height - 300  # Reducido para menos espacio vacío
    doc.setFont("Helvetica-Bold", 9)  # Reducido para ahorrar espacio
    
    # Encabezados de tabla con saltos de línea para ahorrar espacio
    headers = ["CONCEPTO", "Porcentaje\nde recargo", "VALOR\nHORA\nBASE", "HORAS\nLABORADAS", 
               "VALOR\nRECARGO", "VALOR HORA\nCON RECARGO", "VALOR\nTOTAL"]
    # Aumentar márgenes de la tabla para mejor visualización
    x_positions = [30, 130, 230, 330, 430, 480, 530]
    
    for i, header in enumerate(headers):
        # Dibujar encabezado con saltos de línea
        lines = header.split('\n')
        for j, line in enumerate(lines):
            doc.drawString(x_positions[i], y - (j * 10), line)
    
    y -= 30  # Reducido para menos espacio vacío
    doc.setFont("Helvetica", 8)  # Reducido para ahorrar espacio
    
    # === DATOS DE LA TABLA ===
    # Usar la función calcular_valores para obtener los datos correctos
    detalle, resumen = calcular_valores(emp[3], emp[4], horas_dict, config, emp[5] if len(emp) > 5 else 0)
    
    # Dibujar cada concepto en la tabla
    for concepto, porcentaje, valor_hora_base, horas, valor_recargo, subtotal in detalle:
        # Concepto
        doc.drawString(30, y, concepto)
        # Porcentaje de recargo
        doc.drawString(130, y, f"{porcentaje:.2f}")
        # Valor hora base
        doc.drawString(230, y, f"${valor_hora_base:,.0f}")
        # Horas laboradas
        doc.drawString(330, y, f"{horas:.2f}")
        # Valor del recargo
        doc.drawString(430, y, f"${valor_recargo:,.0f}")
        # Valor hora con recargo
        doc.drawString(480, y, f"${valor_hora_base + valor_recargo:,.0f}")
        # Valor total
        doc.drawString(530, y, f"${subtotal:,.0f}")
        
        y -= 18  # Reducido para menos espacio vacío
        
        # Si no hay más espacio, crear nueva página
        if y < 100:
            doc.showPage()
            y = height - 50
            doc.setFont("Helvetica", 8)
    
    # === TOTALES ===
    # Alinear con los subtítulos de arriba, no centrado
    y -= 20
    doc.setFont("Helvetica-Bold", 10)
    
    # Total Horas
    doc.drawString(30, y, f"Total Horas: {total_horas:.2f}")
    
    # Total a pagar
    doc.drawString(30, y - 20, f"TOTAL A PAGAR: ${resumen['total']:,.0f}")
    
    # Neto a pagar (después de deducciones)
    doc.drawString(30, y - 40, f"NETO A PAGAR: ${resumen['neto']:,.0f}")
    
    # === FIRMA ===
    # Agregar salto de línea para más espacio
    y -= 60
    doc.drawString(200, y, "Firma")
    y -= 20
    doc.drawString(200, y, "Prestador del servicio")
    
    # Líneas para firma
    y -= 30
    doc.line(150, y, 350, y)  # Línea para firma
    y -= 20
    doc.line(150, y, 350, y)  # Línea para valor
    
    y -= 30  # Reducido de 35 a 30 para menos espacio vacío
    doc.setFont("Helvetica", 8)  # Reducido de 9 a 8 para ahorrar espacio
    
    # === DATOS DE LA TABLA ===
    # Usar el detalle que ya fue calculado y pasado como parámetro
    # No necesitamos recalcular nada aquí
    
    # Agregar conceptos adicionales si existen
    if resumen.get('auxilio', 0) > 0:
        doc.drawString(30, y, "Auxilio de Transporte")
        doc.drawString(130, y, "0.00")
        doc.drawString(230, y, f"${config['auxilio_transporte']:,.0f}")
        doc.drawString(330, y, "0")
        doc.drawString(430, y, "$0")
        doc.drawString(480, y, f"${config['auxilio_transporte']:,.0f}")
        doc.drawString(530, y, f"${resumen['auxilio']:,.0f}")
        y -= 18
    
    if resumen.get('salud', 0) > 0:
        doc.drawString(30, y, "Deducción Salud")
        doc.drawString(130, y, f"{config['porcentajes']['salud']:.2f}")
        doc.drawString(230, y, f"${resumen['salud']:,.0f}")
        doc.drawString(330, y, "0")
        doc.drawString(430, y, "$0")
        doc.drawString(480, y, f"${resumen['salud']:,.0f}")
        doc.drawString(530, y, f"${resumen['salud']:,.0f}")
        y -= 18
    
    if resumen.get('pension', 0) > 0:
        doc.drawString(30, y, "Deducción Pensión")
        doc.drawString(130, y, f"{config['porcentajes']['pension']:.2f}")
        doc.drawString(230, y, f"${resumen['pension']:,.0f}")
        doc.drawString(330, y, "0")
        doc.drawString(430, y, "$0")
        doc.drawString(480, y, f"${resumen['pension']:,.0f}")
        doc.drawString(530, y, f"${resumen['pension']:,.0f}")
        y -= 18
    
    # === TOTALES ===
    y -= 10
    doc.setFont("Helvetica-Bold", 12)
    # Alinear con la tabla, no centrado - usar la misma posición que la columna "HORAS LABORADAS"
    doc.drawString(230, y, f"Total Horas: {total_horas}")
    y -= 20
    
    doc.drawString(230, y, f"TOTAL A PAGAR: $ {resumen['total']:,.0f}")
    y -= 20
    
    if resumen.get('auxilio', 0) > 0:
        doc.drawString(230, y, f"TOTAL (CON AUXILIO): $ {resumen['total'] + resumen['auxilio']:,.0f}")
        y -= 20
    
    doc.drawString(230, y, f"TOTAL NETO A PAGAR: $ {resumen['neto']:,.0f}")
    
    # === SECCIÓN DE FIRMAS ===
    y -= 40
    doc.setFont("Helvetica", 12)
    doc.drawString(100, y, "FIRMA DEL TRABAJADOR:")
    y -= 25  # Salto de línea para más espacio
    doc.drawString(100, y, "__________________________\n")
    y -= 25  # Salto de línea para más espacio
    doc.drawString(100, y, "CÉDULA:")
    y -= 25  # Salto de línea para más espacio
    doc.drawString(100, y, "__________________________")

def generar_pdf_temporal(doc, emp, quincena, detalle, resumen, config, horas_dict, valor_deuda):
    """Generar PDF para empleados temporales con estructura específica."""
    width, height = letter
    
    # === ENCABEZADO PRINCIPAL ===
    doc.setFont("Helvetica-Bold", 18)
    doc.drawString(200, height - 50, "ORDEN DE TRABAJO")
    
    # === INFORMACIÓN DEL EMPLEADO ===
    doc.setFont("Helvetica-Bold", 12)
    doc.drawString(100, height - 80, f"EMPLEADO: {emp[1].upper()}")
    doc.drawString(100, height - 100, f"CÉDULA: {emp[2]}")
    
    # === PERIODO DE TRABAJO ===
    # La fecha seleccionada solo afecta fecha inicio y fecha fin
    # No es un comprobante por día, es por quincena
    try:
        if "_" in quincena:
            partes = quincena.split("_")
            if len(partes) == 3:
                dia, mes, año = partes
                # Para quincenas: del 1 al 15 o del 16 al 31
                if int(dia) <= 15:
                    fecha_inicio = f"01/{mes}/{año}"
                    fecha_fin = f"15/{mes}/{año}"
                else:
                    fecha_inicio = f"16/{mes}/{año}"
                    fecha_fin = f"31/{mes}/{año}"
            else:
                fecha_inicio = fecha_fin = quincena
        else:
            fecha_inicio = fecha_fin = quincena
    except:
        fecha_inicio = fecha_fin = quincena
    
    doc.drawString(100, height - 120, f"FECHA DE INICIO: {fecha_inicio}")
    doc.drawString(100, height - 140, f"FECHA DE TERMINACIÓN: {fecha_fin}")
    
    # === HORAS Y SALARIO ===
    total_horas = sum(horas_dict.values())
    doc.drawString(100, height - 160, f"Horas prestación de servicio: {total_horas}")
    doc.drawString(100, height - 180, f"SALARIO BASE LIQUIDACIÓN: $ {emp[4]:,.0f}")
    
    # === TABLA PRINCIPAL DE CONCEPTOS ===
    y = height - 230  # Reducido de 210 a 230 para menos espacio vacío
    doc.setFont("Helvetica-Bold", 9)  # Reducido de 10 a 9 para ahorrar espacio
    
    # Encabezados de tabla con saltos de línea para ahorrar espacio
    headers = ["Porcentaje\nde recargo", "VALOR\nHORA\nBASE", "HORAS\nLABORADAS", 
                "VALOR\nRECARGO", "VALOR HORA\nCON RECARGO", "VALOR\nTOTAL"]
    # Aumentar márgenes de la tabla para mejor visualización
    x_positions = [30, 130, 230, 330, 430, 480]
    
    for i, header in enumerate(headers):
        # Dibujar encabezado con saltos de línea
        lines = header.split('\n')
        for j, line in enumerate(lines):
            doc.drawString(x_positions[i], y - (j * 10), line)  # Reducido de 12 a 10
    
    y -= 30  # Reducido de 35 a 30 para menos espacio vacío
    doc.setFont("Helvetica", 8)  # Reducido de 9 a 8 para ahorrar espacio
    
    # === DATOS DE LA TABLA ===
    # IMPORTANTE: Siempre usar el salario mínimo estándar dividido por 220
    valor_hora_base = config["salario_minimo"] / 220
    
    # Solo mostrar conceptos que tengan horas > 0
    conceptos_con_horas = []
    
    if horas_dict["ordinarias"] > 0:
        conceptos_con_horas.append(("Salario Devengado", 0.00, valor_hora_base, horas_dict["ordinarias"], 0, valor_hora_base * horas_dict["ordinarias"]))
    
    if horas_dict["recargo_nocturno"] > 0:
        conceptos_con_horas.append(("Recargo Nocturno", 0.35, valor_hora_base, horas_dict["recargo_nocturno"], valor_hora_base * 0.35, valor_hora_base * 1.35 * horas_dict["recargo_nocturno"]))
    
    if horas_dict["recargo_diurno_dominical"] > 0:
        conceptos_con_horas.append(("Recargo dominical", 0.75, valor_hora_base, horas_dict["recargo_diurno_dominical"], valor_hora_base * 0.75, valor_hora_base * 1.75 * horas_dict["recargo_diurno_dominical"]))
    
    if horas_dict["recargo_nocturno_dominical"] > 0:
        conceptos_con_horas.append(("Recargo Nocturno dominical", 1.10, valor_hora_base, horas_dict["recargo_nocturno_dominical"], valor_hora_base * 1.10, valor_hora_base * 2.10 * horas_dict["recargo_nocturno_dominical"]))
    
    if horas_dict["hora_extra_diurna"] > 0:
        conceptos_con_horas.append(("Hora extra diurna", 0.25, valor_hora_base, horas_dict["hora_extra_diurna"], valor_hora_base * 0.25, valor_hora_base * 1.25 * horas_dict["hora_extra_diurna"]))
    
    if horas_dict["hora_extra_nocturna"] > 0:
        conceptos_con_horas.append(("Hora extra nocturna", 0.75, valor_hora_base, horas_dict["hora_extra_nocturna"], valor_hora_base * 0.75, valor_hora_base * 1.75 * horas_dict["hora_extra_nocturna"]))
    
    if horas_dict["hora_diurna_dominical_o_festivo"] > 0:
        conceptos_con_horas.append(("Hora diurna dominical", 0.80, valor_hora_base, horas_dict["hora_diurna_dominical_o_festivo"], valor_hora_base * 0.80, valor_hora_base * 1.80 * horas_dict["hora_diurna_dominical_o_festivo"]))
    
    if horas_dict["hora_extra_diurna_dominical_o_festivo"] > 0:
        conceptos_con_horas.append(("Hora extra diurna dominical", 1.10, valor_hora_base, horas_dict["hora_extra_diurna_dominical_o_festivo"], valor_hora_base * 1.10, valor_hora_base * 2.10 * horas_dict["hora_extra_diurna_dominical_o_festivo"]))
    
    if horas_dict["hora_nocturna_dominical_o_festivo"] > 0:
        conceptos_con_horas.append(("Hora nocturna dominical", 1.05, valor_hora_base, horas_dict["hora_nocturna_dominical_o_festivo"], valor_hora_base * 1.05, valor_hora_base * 2.05 * horas_dict["hora_nocturna_dominical_o_festivo"]))
    
    # Mostrar conceptos con horas > 0
    for concepto, porcentaje, valor_hora_base, horas, valor_recargo, subtotal in conceptos_con_horas:
        # Concepto (alineado a la izquierda)
        doc.drawString(50, y, f"{porcentaje:.2f}")
        # Valores (alineados a la derecha para mejor legibilidad)
        doc.drawString(120, y, f"${valor_hora_base:,.0f}")
        doc.drawString(200, y, f"{horas}")
        doc.drawString(280, y, f"${valor_recargo:,.0f}")
        doc.drawString(360, y, f"${valor_hora_base + valor_recargo:,.0f}")
        doc.drawString(450, y, f"${subtotal:,.0f}")
        y -= 20  # Aumentado de 18 a 20 para mejor espaciado
    
    # === TOTALES ===
    y -= 10
    doc.setFont("Helvetica-Bold", 12)
    # Alinear con la tabla, no centrado
    doc.drawString(200, y, f"Total Horas: {total_horas}")
    y -= 20
    
    doc.drawString(200, y, f"TOTAL A PAGAR: $ {resumen['total']:,.0f}")
    
    # === SECCIÓN DE FIRMAS ===
    y -= 40
    doc.setFont("Helvetica", 12)
    doc.drawString(100, y, "FIRMA DEL TRABAJADOR: __________________________")
    doc.drawString(100, y - 20, "CÉDULA: __________________________")

# === FUNCIONES MEJORADAS CON PLANTILLAS ===
def exportar_cuadre_caja_plantilla(fecha: str | None = None):
    """
    Exportar cuadre de caja usando plantilla existente CUADRE_CAJA.xlsx.
    
    Args:
        fecha: Fecha en formato "DD-MM-YYYY" o "YYYY-MM-DD"
        
    Returns:
        str: Ruta del archivo generado
    """
    try:
        # Usar plantilla existente
        template_path = os.path.join(EXCEL_DIR, "CUADRE_CAJA.xlsx")
        
        # Verificar que exista la plantilla y tenga los encabezados correctos
        if not os.path.exists(template_path):
            # Crear plantilla básica si no existe
            # Creando plantilla básica
            wb = Workbook()
            ws = wb.active
            if ws:
                ws.title = "Cuadre de Caja"
                
                # Configurar encabezados - usar los que realmente están en la plantilla
                ws['B6'] = "FECHA"
                ws['A9'] = "FACTURA"  # Cambiado de "NÚMERO FACTURA" a "FACTURA"
                ws['B9'] = "AREA"     # Cambiado de "ÁREA" a "AREA"
                ws['C9'] = "FECHA"
                ws['D9'] = "VENTA"
                ws['E9'] = "MODO DE PAGO"
                ws['I10'] = "PROVEEDOR"
                ws['J10'] = "CONCEPTO"
                ws['L10'] = "VALOR"
                
                # Guardar plantilla
                wb.save(template_path)
                # Plantilla básica creada
            else:
                raise Exception("No se pudo crear la hoja de trabajo")
        else:
            # Usando plantilla existente
            
            # Verificar que la plantilla tenga los encabezados correctos
            try:
                wb_check = load_workbook(template_path)
                ws_check = wb_check.active
                
                if ws_check is None:
                    raise Exception("No se pudo abrir la hoja de trabajo de la plantilla")
                
                # Verificar si los encabezados están en su lugar
                if ws_check['B6'].value != "FECHA":
                    # Plantilla existente no tiene encabezado FECHA en B6, actualizando
                    ws_check['B6'] = "FECHA"
                    wb_check.save(template_path)
                    # Plantilla actualizada con encabezados correctos
                
                wb_check.close()
            except Exception as e:
                # Error verificando plantilla
                # Si hay error, crear nueva plantilla
                # Creando nueva plantilla
                wb_new = Workbook()
                ws_new = wb_new.active
                
                if ws_new is None:
                    raise Exception("No se pudo crear la hoja de trabajo")
                
                ws_new.title = "Cuadre de Caja"
                
                # Configurar encabezados
                ws_new['B6'] = "FECHA"
                ws_new['A9'] = "FACTURA"
                ws_new['B9'] = "AREA"
                ws_new['C9'] = "FECHA"
                ws_new['D9'] = "VENTA"
                ws_new['E9'] = "MODO DE PAGO"
                ws_new['I10'] = "PROVEEDOR"
                ws_new['J10'] = "CONCEPTO"
                ws_new['L10'] = "VALOR"
                
                # Guardar nueva plantilla
                wb_new.save(template_path)
                # Nueva plantilla creada
                wb_new.close()
        
        # Normalizar formato de fecha para la base de datos
        fecha_db = ""
        fecha_excel = ""
        
        if fecha:
            # La fecha viene del frontend en formato DD-MM-YYYY
            # Convertir a YYYY-MM-DD para la base de datos
            try:
                from datetime import datetime
                fecha_obj = datetime.strptime(fecha, "%d-%m-%Y")
                fecha_db = fecha_obj.strftime("%Y-%m-%d")
                fecha_excel = fecha  # Mantener formato DD-MM-YYYY para Excel
            except ValueError:
                # Error al parsear fecha
                # Si falla el parsing, usar la fecha tal como viene
                fecha_db = str(fecha) if fecha else ""
                fecha_excel = str(fecha) if fecha else ""
        else:
            # Si no hay fecha, usar la fecha actual
            from datetime import datetime
            colombia_tz = pytz.timezone('America/Bogota')
            fecha_actual = datetime.now(colombia_tz)
            fecha_db = fecha_actual.strftime("%Y-%m-%d")
            fecha_excel = fecha_actual.strftime("%d-%m-%Y")
        
        # Fechas procesadas
        
        # Crear archivo de salida
        output_filename = f"CUADRE_DE_CAJA_{fecha_excel}.xlsx"
        output_path = os.path.join(EXCEL_DIR, output_filename)
        
        # Copiando plantilla
        shutil.copy2(template_path, output_path)
        
        # Abrir y llenar datos
        # Abriendo archivo para llenar datos
        wb = load_workbook(output_path)
        ws = wb.active
        
        if ws is None:
            raise Exception("No se pudo abrir la hoja de trabajo de la plantilla")
        
        # Hoja de trabajo abierta correctamente
        
        # Llenar fecha en B6 (formato DD-MM-YYYY para Excel)
        # IMPORTANTE: Usar la fecha original del usuario, no la fecha actual
        if fecha:
            # Si el usuario seleccionó una fecha específica, usar esa
            ws['B6'] = fecha_excel
        else:
            # Solo si no hay fecha seleccionada, usar la fecha actual
            ws['B6'] = fecha_excel
        
        # Obtener datos de la base de datos usando fecha_db (formato YYYY-MM-DD)
        conn = get_db()
        cursor = conn.cursor()
        
        # Obtener ventas del día específico (NO acumular con otros días)
        
        cursor.execute("""
            SELECT numero_factura, area, monto, metodo_pago, encargado
            FROM cuadre_caja 
            WHERE fecha = ?
            ORDER BY numero_factura
        """, (fecha_db,))
        
        ventas = cursor.fetchall()
        
        # Obtener gastos del día específico (NO acumular con otros días)
        cursor.execute("""
            SELECT proveedor, concepto, valor
            FROM gastos 
            WHERE fecha = ?
            ORDER BY concepto
        """, (fecha_db,))
        
        gastos = cursor.fetchall()
        
        conn.close()
        
        # B5: Encargado (tomar el primero del día o dejar vacío)
        encargado = ""
        if ventas:
            encargado = ventas[0][4] if len(ventas[0]) > 4 and ventas[0][4] else ""
        ws.cell(row=5, column=2, value=encargado)
        
        # Llenar ventas desde la fila 10 (A10, B10, C10, D10, E10)
        fila_venta = 10
        
        for venta in ventas:
            numero_factura, area, monto, metodo_pago, encargado_venta = venta
            
            # A10: Número de factura (columna 1, fila 10)
            ws.cell(row=fila_venta, column=1, value=numero_factura)
            # B10: Área (columna 2, fila 10) - formatear nombre del área
            area_formateada = formatear_nombre_area(area)
            ws.cell(row=fila_venta, column=2, value=area_formateada)
            # C10: Fecha (columna 3, fila 10) - escribir fecha en cada fila
            ws.cell(row=fila_venta, column=3, value=fecha_excel)
            # D10: Venta (monto) (columna 4, fila 10)
            ws.cell(row=fila_venta, column=4, value=monto)
            # E10: Modo de pago (columna 5, fila 10)
            ws.cell(row=fila_venta, column=5, value=metodo_pago)
            
            fila_venta += 1
        
        # Llenar gastos desde la fila 11 (I11, J11, L11)
        fila_gasto = 11
        
        for gasto in gastos:
            proveedor, concepto, valor = gasto
            
            # I11: Proveedor (columna 9, fila 11)
            ws.cell(row=fila_gasto, column=9, value=str(proveedor) if proveedor else "")
            # J11: Concepto (columna 10, fila 11) - está combinada con K11
            ws.cell(row=fila_gasto, column=10, value=str(concepto) if concepto else "")
            # L11: Valor del gasto (columna 12, fila 11)
            ws.cell(row=fila_gasto, column=12, value=valor)
            
            fila_gasto += 1
        
        # Guardar archivo con manejo de errores
        try:
            wb.save(output_path)
            # Cuadre de caja generado
        except PermissionError as e:
            if "being used by another process" in str(e):
                # Si el archivo está en uso, crear uno con nombre único
                import time
                timestamp = int(time.time())
                output_path = os.path.join(EXCEL_DIR, f"CUADRE_DE_CAJA_{timestamp}.xlsx")
                wb.save(output_path)
                # Cuadre de caja generado con nombre único
            else:
                raise e
        
        return output_path
        
    except Exception as e:
        # Error en exportar_cuadre_caja_plantilla
        raise e

def exportar_nomina_plantilla(quincena: str | None = None, tipo_nomina: str = "general"):
    """
    Exportar nómina usando plantilla unificada NOMINA.xlsx.
    
    Args:
        quincena: Quincena en formato "15_MM_YYYY" o "31_MM_YYYY"
        tipo_nomina: "general" o "agosto"
        
    Returns:
        str: Ruta del archivo generado
    """
    try:
        # Usar plantilla unificada NOMINA.xlsx
        template_path = os.path.join(EXCEL_DIR, "NOMINA.xlsx")
        output_filename = "NOMINA_GENERADA.xlsx"
        
        # Verificar que exista la plantilla
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"No se encontró la plantilla: {template_path}")
        
        # Obtener datos de empleados
        conn = get_db()
        cursor = conn.cursor()
        
        if quincena is None:
            # Usar quincena más reciente
            cursor.execute("SELECT DISTINCT quincena FROM horas_trabajadas ORDER BY quincena DESC LIMIT 1")
            result = cursor.fetchone()
            quincena = result[0] if result else "15_08_2025"
        
        # Obtener todos los empleados (fijos y temporales)
        cursor.execute("""
            SELECT e.nombre, e.cedula, e.tipo, e.salario, 
                   h.horas_ordinarias, h.recargo_nocturno, h.recargo_diurno_dominical,
                   h.hora_extra_diurna, h.hora_extra_nocturna,
                   h.hora_diurna_dominical_o_festivo, h.hora_extra_diurna_dominical_o_festivo,
                   h.hora_nocturna_dominical_o_festivo, h.hora_extra_nocturna_dominical_o_festivo,
                   h.motivo_deuda, h.valor_deuda
            FROM empleados e
            LEFT JOIN horas_trabajadas h ON e.id = h.empleado_id AND h.quincena = ?
            ORDER BY e.nombre
        """, (quincena,))
        
        all_employees = cursor.fetchall()
        conn.close()
        
        # Crear archivo de salida
        output_path = os.path.join(EXCEL_DIR, output_filename)
        shutil.copy2(template_path, output_path)
        
        # Abrir y llenar datos
        wb = load_workbook(output_path)
        ws = wb.active
        
        if ws is None:
            raise Exception("No se pudo abrir la hoja de trabajo de la plantilla")
        
        # Buscar empleados en la plantilla y llenar sus datos
        # Empezar desde la fila 8 (como especificó el usuario)
        fila_inicio = 8
        
        for emp in all_employees:
            nombre, cedula, tipo, salario = emp[0], emp[1], emp[2], emp[3]
            horas_dict = {
                "ordinarias": emp[4] or 0,
                "recargo_nocturno": emp[5] or 0,
                "recargo_diurno_dominical": emp[6] or 0,
                "hora_extra_diurna": emp[7] or 0,
                "hora_extra_nocturna": emp[8] or 0,
                "hora_diurna_dominical_o_festivo": emp[9] or 0,
                "hora_extra_diurna_dominical_o_festivo": emp[10] or 0,
                "hora_nocturna_dominical_o_festivo": emp[11] or 0,
                "hora_extra_nocturna_dominical_o_festivo": emp[12] or 0
            }
            
            config = cargar_configuracion()
            detalle, resumen = calcular_valores(tipo, salario, horas_dict, config, emp[14] or 0)
            
            # Buscar la fila donde está el empleado en la plantilla por nombre o cédula
            fila_empleado = None
            
            # Primero buscar por cédula (más preciso)
            if ws.max_row is not None:
                for fila in range(fila_inicio, ws.max_row + 1):
                    cedula_celda = ws.cell(row=fila, column=3).value  # Columna C (cédula)
                    if cedula_celda and str(cedula) == str(cedula_celda):
                        fila_empleado = fila
                        break
                
                # Si no se encuentra por cédula, buscar por nombre
                if not fila_empleado:
                    for fila in range(fila_inicio, ws.max_row + 1):
                        nombre_celda = ws.cell(row=fila, column=2).value  # Columna B (nombre)
                        if nombre_celda and nombre.upper() in str(nombre_celda).upper():
                            fila_empleado = fila
                            break
            
            # Si el empleado no se encuentra, usar espacios vacíos para nuevos usuarios (B8-B18)
            if not fila_empleado:
                for fila in range(8, 18):  # B8 a B18
                    if ws.cell(row=fila, column=2).value is None or str(ws.cell(row=fila, column=2).value).strip() == "":
                        fila_empleado = fila
                        # Llenar nombre y cédula para nuevo usuario
                        ws.cell(row=fila_empleado, column=2, value=nombre)  # B: Nombre
                        ws.cell(row=fila_empleado, column=3, value=cedula)  # C: Cédula
                        break
            
            if fila_empleado:
                # Extraer valores del detalle para mapear correctamente
                # El detalle contiene tuplas: (concepto, porcentaje, valor_hora_base, horas, valor_recargo, subtotal)
                valores_por_concepto = {}
                for concepto, porcentaje, valor_hora_base, horas, valor_recargo, subtotal in detalle:
                    if concepto == "Salario Devengado":
                        valores_por_concepto['salario_devengado'] = subtotal
                    elif concepto == "Hora extra diurna":
                        valores_por_concepto['horas_extras_diurnas'] = subtotal
                    elif concepto == "Recargo Nocturno dominical o festivo":
                        valores_por_concepto['recargos_nocturnos_dominical'] = subtotal
                    elif concepto == "Hora extra nocturna":
                        valores_por_concepto['extra_nocturna'] = subtotal
                    elif concepto == "Recargo Nocturno":
                        valores_por_concepto['recargo_nocturno'] = subtotal
                    elif concepto == "Recargo dominical laborado":
                        valores_por_concepto['recargo_diurno_dominical'] = subtotal
                    elif concepto == "Hora diurna dominical o festivo":
                        valores_por_concepto['hora_diurna_dominical'] = subtotal
                    elif concepto == "Hora extra diurna dominical o festivo":
                        valores_por_concepto['hora_extra_diurna_dominical'] = subtotal
                    elif concepto == "Hora nocturna Dominical o festiva":
                        valores_por_concepto['hora_nocturna_dominical'] = subtotal
                    elif concepto == "Hora extra nocturna dominical o festivo":
                        valores_por_concepto['hora_extra_nocturna_dominical'] = subtotal
                
                # Para empleados fijos, el salario devengado debe ser salario mínimo / 2
                if tipo == 'FIJO':
                    # Obtener salario mínimo de la configuración
                    salario_minimo = config.get('salario_minimo', 1300000)  # Valor por defecto
                    salario_devengado_fijo = salario_minimo / 2
                    # NO sobrescribir aquí, usar el valor del detalle que ya está calculado correctamente
                    
                    # Verificar si el valor del detalle es correcto
                    if valores_por_concepto.get('salario_devengado', 0) != salario_devengado_fijo:
                        # Si no coincide, usar el valor esperado
                        valores_por_concepto['salario_devengado'] = salario_devengado_fijo
                
                # Verificar que el salario devengado esté presente
                if 'salario_devengado' not in valores_por_concepto:
                    # Buscar en el detalle
                    for concepto, porcentaje, valor_hora_base, horas, valor_recargo, subtotal in detalle:
                        if concepto == "Salario Devengado":
                            valores_por_concepto['salario_devengado'] = subtotal
                            break
                
                # Llenar datos según el mapeo correcto especificado por el usuario:
                # B12: Nombre (ya está lleno)
                # C12: Cédula (ya está lleno)
                # D: Salario básico (ya está en la plantilla)
                
                # E: Salario devengado (Horas ordinarias) - usar el valor del detalle
                salario_devengado = valores_por_concepto.get('salario_devengado', 0)
                ws.cell(row=fila_empleado, column=5, value=salario_devengado)
                
                # F: Horas extra diurnas
                horas_extras = valores_por_concepto.get('horas_extras_diurnas', 0)
                ws.cell(row=fila_empleado, column=6, value=horas_extras)
                
                # G: Recargos nocturnos dominicales
                recargos_nocturnos = valores_por_concepto.get('recargos_nocturnos_dominical', 0)
                ws.cell(row=fila_empleado, column=7, value=recargos_nocturnos)
                
                # H: Hora extra nocturna
                extra_nocturna = valores_por_concepto.get('extra_nocturna', 0)
                ws.cell(row=fila_empleado, column=8, value=extra_nocturna)
                
                # I: Recargo nocturno
                recargo_nocturno = valores_por_concepto.get('recargo_nocturno', 0)
                ws.cell(row=fila_empleado, column=9, value=recargo_nocturno)
                
                # J: Recargo dominical laborado
                recargo_dominical = valores_por_concepto.get('recargo_diurno_dominical', 0)
                ws.cell(row=fila_empleado, column=10, value=recargo_dominical)
                
                # K: Hora diurna dominical o festiva
                hora_diurna_dominical = valores_por_concepto.get('hora_diurna_dominical', 0)
                ws.cell(row=fila_empleado, column=11, value=hora_diurna_dominical)
                
                # L: Hora nocturna dominical o festiva
                hora_nocturna_dominical = valores_por_concepto.get('hora_nocturna_dominical', 0)
                ws.cell(row=fila_empleado, column=12, value=hora_nocturna_dominical)
                
                # M: Extra diurna dominical o festiva
                extra_diurna_dominical = valores_por_concepto.get('hora_extra_diurna_dominical', 0)
                ws.cell(row=fila_empleado, column=13, value=extra_diurna_dominical)
                
                # N: Extra nocturna dominical o festiva
                extra_nocturna_dominical = valores_por_concepto.get('hora_extra_nocturna_dominical', 0)
                ws.cell(row=fila_empleado, column=14, value=extra_nocturna_dominical)
                
                # O: Auxilio de transporte (solo para fijos, 100,000) - NO ESCRIBIR (fórmula de Excel)
                # P: Total devengado - NO ESCRIBIR (fórmula de Excel)
                # Q: Salud - NO ESCRIBIR (fórmula de Excel)
                # R: Pensión - NO ESCRIBIR (fórmula de Excel)
                # S: Neto pagado - NO ESCRIBIR (fórmula de Excel)
                
                # Para empleados temporales, no mostrar auxilio de transporte
                if tipo == 'TEMPORAL':
                    # O: Auxilio de transporte (0 para temporales)
                    ws.cell(row=fila_empleado, column=15, value=0)
                else:
                    # Para empleados fijos, aplicar la fórmula del auxilio de transporte
                    # Fórmula: =SI(D8<=1423500*2; (200000/30)*(F8/(D8/30)); 0)
                    # Donde D8 es salario básico, F8 es salario devengado
                    # Solo para empleados fijos, sin importar el salario
                    salario_basico = salario
                    salario_devengado = valores_por_concepto.get('salario_devengado', 0)
                    
                    # Para empleados fijos: SIEMPRE auxilio de transporte
                    if tipo == 'FIJO':
                        # Calcular auxilio de transporte
                        auxilio_transporte = (200000 / 30) * (salario_devengado / (salario_basico / 30))
                        ws.cell(row=fila_empleado, column=15, value=auxilio_transporte)
                    else:
                        # Para empleados temporales: SIN auxilio de transporte
                        ws.cell(row=fila_empleado, column=15, value=0)
                
                # Aplicar fórmulas de Excel para salud y pensión
                # Salud = E8 * 4% (donde E8 es salario devengado)
                salud = salario_devengado * 0.04
                ws.cell(row=fila_empleado, column=17, value=salud)  # Columna Q
                
                # Pensión = E8 * 4% (donde E8 es salario devengado)
                pension = salario_devengado * 0.04
                ws.cell(row=fila_empleado, column=18, value=pension)  # Columna R
        
        # Guardar archivo con manejo de errores
        try:
            wb.save(output_path)
            # Nómina generada
        except PermissionError as e:
            if "being used by another process" in str(e):
                # Si el archivo está en uso, crear uno con nombre único
                import time
                timestamp = int(time.time())
                output_path = os.path.join(EXCEL_DIR, f"NOMINA_GENERADA_{timestamp}.xlsx")
                wb.save(output_path)
                # Nómina generada con nombre único
            else:
                raise e
        
        return output_path
        
    except Exception as e:
        # Error en exportar_nomina_plantilla
        raise e

def guardar_cuadre_caja(datos):
    """
    Guardar datos de cuadre de caja en la base de datos.
    
    Args:
        datos: Diccionario con los datos a guardar
            - fecha: str (formato YYYY-MM-DD)
            - numero_factura: str
            - area: str
            - monto: float
            - metodo_pago: str
    """
    conn = None
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute("""
            INSERT INTO cuadre_caja (fecha, numero_factura, area, monto, metodo_pago, encargado)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            datos["fecha"],
            datos["numero_factura"],
            datos["area"],
            datos["monto"],
            datos["metodo_pago"],
            datos.get("encargado", "")  # Campo opcional, por defecto vacío
        ))
        
        conn.commit()
        # Cuadre de caja guardado
        
    except Exception as e:
        # Error al guardar cuadre de caja
        raise e
    finally:
        if conn:
            conn.close()

# === FUNCIONES AUXILIARES ===
def obtener_estadisticas():
    """Obtener estadísticas generales del sistema."""
    conn = get_db()
    cursor = conn.cursor()
    
    # Total empleados
    cursor.execute("SELECT COUNT(*) FROM empleados")
    total_empleados = cursor.fetchone()[0]
    
    # Empleados por tipo
    cursor.execute("SELECT tipo, COUNT(*) FROM empleados GROUP BY tipo")
    empleados_por_tipo = dict(cursor.fetchall())
    
    # Total horas registradas
    cursor.execute("SELECT COUNT(*) FROM horas_trabajadas")
    total_horas = cursor.fetchone()[0]
    
    # Total cuadres de caja
    cursor.execute("SELECT COUNT(*) FROM cuadre_caja")
    total_cuadres = cursor.fetchone()[0]
    
    conn.close()
    
    return {
        "total_empleados": total_empleados,
        "empleados_por_tipo": empleados_por_tipo,
        "total_horas": total_horas,
        "total_cuadres": total_cuadres
    }

def obtener_empleado_por_id(empleado_id):
    """Obtener un empleado específico por su ID."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM empleados WHERE id = ?", (empleado_id,))
    empleado = cursor.fetchone()
    conn.close()
    return empleado

def obtener_horas_empleado(empleado_id, quincena):
    """Obtener las horas trabajadas de un empleado en una quincena específica."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT horas_ordinarias, recargo_nocturno, recargo_diurno_dominical,
               recargo_nocturno_dominical, hora_extra_diurna, hora_extra_nocturna,
               hora_diurna_dominical_o_festivo, hora_extra_diurna_dominical_o_festivo,
               hora_nocturna_dominical_o_festivo, hora_extra_nocturna_dominical_o_festivo,
               motivo_deuda, valor_deuda, descuento_inasistencia, otros_descuentos, valor_otros_descuentos
        FROM horas_trabajadas 
        WHERE empleado_id = ? AND quincena = ?
    """, (empleado_id, quincena))
    horas = cursor.fetchone()
    conn.close()
    
    if horas:
        # Convertir a diccionario para facilitar el uso en el template
        return {
            "horas_ordinarias": horas[0] or 0,
            "recargo_nocturno": horas[1] or 0,
            "recargo_diurno_dominical": horas[2] or 0,
            "recargo_nocturno_dominical": horas[3] or 0,
            "hora_extra_diurna": horas[4] or 0,
            "hora_extra_nocturna": horas[5] or 0,
            "hora_diurna_dominical_o_festivo": horas[6] or 0,
            "hora_extra_diurna_dominical_o_festivo": horas[7] or 0,
            "hora_nocturna_dominical_o_festivo": horas[8] or 0,
            "hora_extra_nocturna_dominical_o_festivo": horas[9] or 0,
            "motivo_deuda": horas[10] or "",
            "valor_deuda": horas[11] or 0,
            "descuento_inasistencia": horas[12] or 0,
            "otros_descuentos": horas[13] or "",
            "valor_otros_descuentos": horas[14] or 0
        }
    return None

def eliminar_empleado(empleado_id):
    """Eliminar un empleado y todas sus horas trabajadas."""
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM horas_trabajadas WHERE empleado_id = ?", (empleado_id,))
        cursor.execute("DELETE FROM empleados WHERE id = ?", (empleado_id,))
        conn.commit()
        return True
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        conn.close()

def obtener_quincena_actual():
    """Obtener la quincena actual en formato YYYY-MM."""
    from datetime import datetime
    return datetime.now().strftime("%Y-%m")

def obtener_quincenas_disponibles():
    """Obtener lista de quincenas disponibles en la base de datos."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT quincena FROM horas_trabajadas ORDER BY quincena DESC")
    quincenas = [row[0] for row in cursor.fetchall()]
    conn.close()
    return quincenas

# === MODELS PYDANTIC ===
class Venta(BaseModel):
    """Modelo para una venta individual."""
    factura: Optional[str] = None
    area: str
    valor: float
    forma_pago: str
    descripcion: Optional[str] = None

class Gasto(BaseModel):
    """Modelo para un gasto individual."""
    descripcion: str
    valor: float

class PayloadCuadre(BaseModel):
    """Payload completo para generar cuadre de caja."""
    fecha: str  # Formato: "DD-MM-YYYY"
    ventas: List[Venta] = []
    gastos: List[Gasto] = []

class PayloadNomina(BaseModel):
    """Payload para generar nómina."""
    quincena: str  # Formato: "15_MM_YYYY" o "31_MM_YYYY"
    fecha_exportacion: str  # Formato: "DD-MM-YYYY"
    tipo_nomina: str = "general"  # "general" o "agosto"

def formatear_nombre_area(area: str) -> str:
    """
    Formatear el nombre del área para mostrar en Excel.
    Convierte nombres como 'dia_de_sol' a 'Dia de Sol'.
    
    Args:
        area: Nombre del área con guiones bajos
        
    Returns:
        str: Nombre del área formateado
    """
    if not area:
        return area
    
    # Reemplazar guiones bajos con espacios
    area_con_espacios = area.replace('_', ' ')
    
    # Capitalizar cada palabra
    palabras = area_con_espacios.split()
    palabras_capitalizadas = [palabra.capitalize() for palabra in palabras]
    
    # Unir las palabras
    return ' '.join(palabras_capitalizadas)

def generar_reporte_general(fecha_inicio: str, fecha_fin: str) -> str:
    """
    Generar reporte general del sistema en formato Excel.
    
    Args:
        fecha_inicio: Fecha de inicio en formato YYYY-MM-DD
        fecha_fin: Fecha de fin en formato YYYY-MM-DD
        
    Returns:
        str: Ruta del archivo Excel generado
    """
    try:
        # Crear un nuevo libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte General"
        
        # Estilos
        titulo_font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
        subtitulo_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        normal_font = Font(name='Arial', size=10)
        
        titulo_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        subtitulo_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        
        # Título principal
        ws['A1'] = "REPORTE GENERAL DEL SISTEMA"
        ws['A1'].font = titulo_font
        ws['A1'].fill = titulo_fill
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Período del reporte
        ws['A2'] = f"Período: {fecha_inicio} a {fecha_fin}"
        ws['A2'].font = subtitulo_font
        ws['A2'].fill = subtitulo_fill
        ws.merge_cells('A2:H2')
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Fecha de generación
        ws['A3'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A3'].font = normal_font
        ws.merge_cells('A3:H3')
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Espacio
        ws['A5'] = ""
        
        # Sección 1: Resumen de Empleados
        ws['A6'] = "RESUMEN DE EMPLEADOS"
        ws['A6'].font = subtitulo_font
        ws['A6'].fill = subtitulo_fill
        ws.merge_cells('A6:H6')
        ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers de empleados
        headers_empleados = ["ID", "Nombre", "Cédula", "Tipo Contrato", "Salario", "Fecha Registro"]
        for col, header in enumerate(headers_empleados, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos de empleados
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, nombre, cedula, tipo, salario, fecha_registro 
            FROM empleados 
            ORDER BY nombre
        """)
        empleados = cursor.fetchall()
        
        for row_idx, empleado in enumerate(empleados, 8):
            for col_idx, valor in enumerate(empleado, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.font = normal_font
                if col_idx == 5:  # Salario
                    cell.number_format = '#,##0'
        
        # Espacio
        ws[f'A{8 + len(empleados)}'] = ""
        
        # Sección 2: Resumen de Ventas por Período
        row_actual = 9 + len(empleados)
        ws[f'A{row_actual}'] = "RESUMEN DE VENTAS POR PERÍODO"
        ws[f'A{row_actual}'].font = subtitulo_font
        ws[f'A{row_actual}'].fill = subtitulo_fill
        ws.merge_cells(f'A{row_actual}:H{row_actual}')
        ws[f'A{row_actual}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers de ventas
        headers_ventas = ["Fecha", "N° Factura", "Área", "Monto", "Método Pago", "Encargado"]
        for col, header in enumerate(headers_ventas, 1):
            cell = ws.cell(row=row_actual + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos de ventas
        cursor.execute("""
            SELECT fecha, numero_factura, area, monto, metodo_pago, encargado
            FROM cuadre_caja 
            WHERE fecha BETWEEN ? AND ?
            ORDER BY fecha, numero_factura
        """, (fecha_inicio, fecha_fin))
        ventas = cursor.fetchall()
        
        for row_idx, venta in enumerate(ventas, row_actual + 2):
            for col_idx, valor in enumerate(venta, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.font = normal_font
                if col_idx == 4:  # Monto
                    cell.number_format = '#,##0'
        
        # Espacio
        row_actual = row_actual + 2 + len(ventas)
        ws[f'A{row_actual}'] = ""
        
        # Sección 3: Resumen de Gastos por Período
        row_actual += 1
        ws[f'A{row_actual}'] = "RESUMEN DE GASTOS POR PERÍODO"
        ws[f'A{row_actual}'].font = subtitulo_font
        ws[f'A{row_actual}'].fill = subtitulo_fill
        ws.merge_cells(f'A{row_actual}:H{row_actual}')
        ws[f'A{row_actual}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers de gastos
        headers_gastos = ["Fecha", "Proveedor", "Concepto", "Valor"]
        for col, header in enumerate(headers_gastos, 1):
            cell = ws.cell(row=row_actual + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos de gastos
        cursor.execute("""
            SELECT fecha, proveedor, concepto, valor
            FROM gastos 
            WHERE fecha BETWEEN ? AND ?
            ORDER BY fecha, concepto
        """, (fecha_inicio, fecha_fin))
        gastos = cursor.fetchall()
        
        for row_idx, gasto in enumerate(gastos, row_actual + 2):
            for col_idx, valor in enumerate(gasto, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.font = normal_font
                if col_idx == 4:  # Valor
                    cell.number_format = '#,##0'
        
        # Espacio
        row_actual = row_actual + 2 + len(gastos)
        ws[f'A{row_actual}'] = ""
        
        # Sección 4: Totales y Estadísticas
        row_actual += 1
        ws[f'A{row_actual}'] = "TOTALES Y ESTADÍSTICAS"
        ws[f'A{row_actual}'].font = subtitulo_font
        ws[f'A{row_actual}'].fill = subtitulo_fill
        ws.merge_cells(f'A{row_actual}:H{row_actual}')
        ws[f'A{row_actual}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Calcular totales
        total_ventas = sum(venta[3] for venta in ventas) if ventas else 0
        total_gastos = sum(gasto[3] for gasto in gastos) if gastos else 0
        balance = total_ventas - total_gastos
        
        # Escribir totales
        row_actual += 1
        ws[f'A{row_actual}'] = "Total Ventas:"
        ws[f'B{row_actual}'] = total_ventas
        ws[f'B{row_actual}'].number_format = '#,##0'
        ws[f'B{row_actual}'].font = Font(bold=True)
        
        row_actual += 1
        ws[f'A{row_actual}'] = "Total Gastos:"
        ws[f'B{row_actual}'] = total_gastos
        ws[f'B{row_actual}'].number_format = '#,##0'
        ws[f'B{row_actual}'].font = Font(bold=True)
        
        row_actual += 1
        ws[f'A{row_actual}'] = "Balance:"
        ws[f'B{row_actual}'] = balance
        ws[f'B{row_actual}'].number_format = '#,##0'
        ws[f'B{row_actual}'].font = Font(bold=True)
        
        # Ajustar ancho de columnas
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Cerrar conexión
        conn.close()
        
        # Guardar archivo
        filename = f"REPORTE_GENERAL_{fecha_inicio}_{fecha_fin}.xlsx"
        filepath = os.path.join(EXCEL_DIR, filename)
        wb.save(filepath)
        
        return filepath
        
    except Exception as e:
        if 'conn' in locals():
            conn.close()
        raise Exception(f"Error al generar reporte general: {str(e)}")
