import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
import json
import os
import sys
import webbrowser
import random
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

def preparar_entorno():
    if not os.path.exists(PDF_DIR):
        os.makedirs(PDF_DIR)

CONFIG_PATH = "ley_colombia_2025.json"
DB_PATH = "empleados.db"
PDF_DIR = "comprobantes"
LOG_PATH = "nomina.log"
RESUMEN_PATH = "resumen_empleados.txt"
HTML_RESUMEN_FIJO = "resumen_empleados_fijos.html"
HTML_RESUMEN_TEMPORAL = "resumen_empleados_temporales.html"

TIPOS_HORAS = [
    ("ordinarias", 0.00),
    ("recargo_nocturno", 0.35),
    ("recargo_diurno_dominical", 0.75),
    ("recargo_nocturno_dominical", 1.10),
    ("hora_extra_diurna", 0.25),
    ("hora_extra_nocturna", 0.75),
    ("hora_diurna_dominical_o_festivo", 0.80),
    ("hora_extra_diurna_dominical_o_festivo", 1.05),
    ("hora_nocturna_dominical_o_festivo", 1.10),
    ("hora_extra_nocturna_dominical_o_festivo", 1.85)
]

def cargar_configuracion():
    default_config = {
        "salario_minimo": 1423500,
        "auxilio_transporte": 100000,
        "descuentos": {"salud": 0.04, "pension": 0.04},
        "empresa": {
            "nombre": "Villa Venecia",
            "nit": "901.234.567-8",
            "direccion": "CRA. 43 SUCRE, VENECIA, ANTIOQUIA, COLOMBIA"
        }
    }
    if not os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(default_config, f, indent=4, ensure_ascii=False)
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def guardar_configuracion(config):
    try:
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=4, ensure_ascii=False)
    except Exception as e:
        raise Exception(f"Error al guardar configuración: {str(e)}")

def crear_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS empleados (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        cedula TEXT NOT NULL UNIQUE,
        tipo TEXT NOT NULL,
        salario REAL NOT NULL,
        comentario TEXT
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS horas_trabajadas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        empleado_id INTEGER,
        quincena TEXT,
        tipo_hora TEXT,
        horas REAL,
        deuda_comentario TEXT,
        deuda_valor REAL,
        FOREIGN KEY (empleado_id) REFERENCES empleados(id)
    )""")
    conn.commit()
    conn.close()

def guardar_empleado(nombre, cedula, tipo, salario, comentario):
    if not all([nombre, cedula, tipo, salario]):
        messagebox.showerror("ERROR", "Todos los campos son obligatorios")
        registrar_log("Intento de guardado fallido: Campos vacíos")
        return False
    try:
        salario = float(salario.replace(',', ''))
        if salario <= 0:
            raise ValueError("El salario debe ser mayor a 0")
    except ValueError as e:
        messagebox.showerror("ERROR", f"El salario debe ser un número válido: {str(e)}")
        registrar_log(f"Intento de guardado fallido: Salario inválido - {str(e)}")
        return False
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute("INSERT INTO empleados (nombre, cedula, tipo, salario, comentario) VALUES (?, ?, ?, ?, ?)",
                    (nombre, cedula, tipo, salario, comentario))
        conn.commit()
        registrar_log(f"Empleado registrado con éxito: {nombre} ({cedula})")
        return True
    except sqlite3.IntegrityError as e:
        messagebox.showerror("ERROR", f"Ya existe un empleado con la cédula {cedula}: {str(e)}")
        registrar_log(f"Intento de guardado fallido: Cédula duplicada - {str(e)}")
        return False
    except Exception as e:
        messagebox.showerror("ERROR", f"Error al guardar el empleado: {str(e)}")
        registrar_log(f"Intento de guardado fallido: Error inesperado - {str(e)}")
        return False
    finally:
        conn.close()

# Obtener empleados
def obtener_empleados():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT id, nombre, cedula, tipo, salario, comentario FROM empleados")
    datos = c.fetchall()
    conn.close()
    return datos

# Guardar horas trabajadas
def guardar_horas_trabajadas(empleado_id, horas_dict, quincena):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute("DELETE FROM horas_trabajadas WHERE empleado_id = ? AND quincena = ?", (empleado_id, quincena))
        for tipo_hora, horas in horas_dict.items():
            if tipo_hora not in ["deuda_comentario", "deuda_valor"] and horas > 0:
                c.execute("INSERT INTO horas_trabajadas (empleado_id, quincena, tipo_hora, horas, deuda_comentario, deuda_valor) VALUES (?, ?, ?, ?, ?, ?)",
                            (empleado_id, quincena, tipo_hora, horas, horas_dict.get("deuda_comentario", ""), float(horas_dict.get("deuda_valor", 0)) if horas_dict.get("deuda_valor") else 0))
        conn.commit()
        registrar_log(f"Horas guardadas para empleado ID {empleado_id} en quincena {quincena}")
    except Exception as e:
        registrar_log(f"Error guardando horas para empleado ID {empleado_id}: {str(e)}")
        messagebox.showerror("ERROR", f"Error al guardar horas: {str(e)}")
    finally:
        conn.close()

# Obtener horas trabajadas
def obtener_horas_trabajadas(empleado_id, quincena):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT tipo_hora, horas, deuda_comentario, deuda_valor FROM horas_trabajadas WHERE empleado_id = ? AND quincena = ?",
                (empleado_id, quincena))
    datos = c.fetchall()
    conn.close()
    horas_dict = {th[0]: 0 for th in TIPOS_HORAS}
    for row in datos:
        horas_dict[row[0]] = row[1]
        if row[2]:
            horas_dict["deuda_comentario"] = row[2]
        if row[3]:
            horas_dict["deuda_valor"] = row[3]
    return horas_dict

# Registrar en log
def registrar_log(mensaje):
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(f"[{datetime.now()}] {mensaje}\n")

# Calcular valores de nómina
def calcular_valores(tipo, salario_base, horas_dict, config):
    valor_hora_base = salario_base / 220
    detalle = []
    total = 0
    total_horas = 0

    # Siempre incluir horas ordinarias
    horas_ordinarias = horas_dict.get("ordinarias", 0)
    
    # Para empleados fijos, el subtotal de horas ordinarias es exactamente la mitad del salario
    if tipo == "FIJO":
        subtotal_ordinarias = round(salario_base / 2, 2)
    else:
        # Para empleados temporales, se calcula normalmente
        subtotal_ordinarias = round(horas_ordinarias * valor_hora_base, 2)
    
    total += subtotal_ordinarias
    total_horas += horas_ordinarias
    detalle.append((
        "Ordinarias",
        round(valor_hora_base, 2),
        horas_ordinarias,
        0.00,
        round(valor_hora_base, 2),
        subtotal_ordinarias
    ))

    # Recorrer todos los tipos de horas con recargo
    for th, recargo in TIPOS_HORAS[1:]:
        horas = horas_dict.get(th, 0)
        valor_recargo = valor_hora_base * recargo
        valor_total_hora = valor_hora_base + valor_recargo
        subtotal = horas * valor_total_hora
        total += subtotal
        total_horas += horas
        detalle.append((
            th.replace("_", " ").capitalize(),
            round(valor_hora_base, 2),
            horas,
            round(valor_recargo, 2),
            round(valor_total_hora, 2),
            round(subtotal, 2)
        ))

    # Cálculos finales según tipo
    if tipo == "FIJO":
        # Para empleados fijos, la base de cotización es exactamente la mitad del salario
        tope_cotizacion = subtotal_ordinarias  # Ya es salario_base / 2
        salud = tope_cotizacion * config["descuentos"]["salud"]
        pension = tope_cotizacion * config["descuentos"]["pension"]
        deuda = float(horas_dict.get("deuda_valor", 0)) if horas_dict.get("deuda_valor") else 0
        auxilio = config["auxilio_transporte"] if salario_base <= 2 * config["salario_minimo"] and horas_ordinarias > 0 else 0
        neto = total + auxilio - salud - pension - deuda
    else:
        salud = 0
        pension = 0
        deuda = float(horas_dict.get("deuda_valor", 0)) if horas_dict.get("deuda_valor") else 0
        auxilio = 0
        neto = total - deuda

    return detalle, {
        "total": round(total, 2),
        "auxilio": round(auxilio, 2),
        "salud": round(salud, 2),
        "pension": round(pension, 2),
        "neto": round(neto, 2),
        "total_horas": round(total_horas, 1)
    }

# Generar PDF
def generar_pdf(nombre, cedula, tipo, salario_base, horas_dict, config, deuda_comentario="", deuda_valor=0):
    if not os.path.exists(PDF_DIR):
        try:
            os.makedirs(PDF_DIR)
        except OSError as e:
            registrar_log(f"Error creando directorio {PDF_DIR}: {str(e)}")
            messagebox.showerror("ERROR", f"No se pudo crear el directorio {PDF_DIR}: {str(e)}")
            return
    nombre_seguro = "".join(c if c.isalnum() or c in " _-" else "_" for c in nombre.upper().replace(" ", "_"))
    fecha = datetime.now().strftime("%m-%d-%Y_%I-%M-%p").replace(":", "-")
    archivo = f"{PDF_DIR}/COMPROBANTE_{nombre_seguro}_{cedula}_{fecha}.pdf"
    try:
        doc = SimpleDocTemplate(archivo, pagesize=letter, rightMargin=36, leftMargin=72, topMargin=72, bottomMargin=72)
        elements = []
        styles = getSampleStyleSheet()

        custom_style = styles["Normal"].clone('CustomStyle')
        custom_style.fontName = "Helvetica"
        custom_style.fontSize = 10
        custom_style.leading = 12
        custom_style.textColor = "#000000"
        custom_style.alignment = 0

        elements.append(Paragraph(f"ORDEN DE TRABAJO N° {fecha.upper()}", styles["Heading1"]))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"EMPRESA: {config['empresa']['nombre'].upper()}", custom_style))
        elements.append(Paragraph(f"NIT: {config['empresa']['nit'].upper()}", custom_style))
        elements.append(Paragraph(f"DIRECCIÓN: {config['empresa']['direccion'].upper()}", custom_style))
        elements.append(Spacer(1, 12))

        elements.append(Paragraph(f"PRESTADOR DEL SERVICIO: {nombre.upper()}", custom_style))
        elements.append(Paragraph(f"CÉDULA: {cedula.upper()}", custom_style))
        elements.append(Paragraph(f"TIPO DE CONTRATO: {tipo.upper()}", custom_style))
        elements.append(Paragraph(f"SALARIO BASE LIQUIDACIÓN: ${salario_base:,.0f}", custom_style))
        elements.append(Paragraph(f"TOTAL HORAS TRABAJADAS: {horas_dict['total_horas']:.1f}", custom_style))
        elements.append(Spacer(1, 12))

        table_data = [
            ["CONCEPTO", "VALOR HORA", "VALOR RECARGO", "VALOR TOTAL", "HORAS", "SUBTOTAL"],
        ]
        detalle, resumen = calcular_valores(tipo, salario_base, horas_dict, config)
        if not detalle:
            raise ValueError("No hay datos de horas para generar el PDF")
        
        # Para empleados fijos, corregir el total de horas ordinarias
        if tipo == "FIJO":
            # Convertir tuplas a listas para poder modificarlas
            detalle_modificable = [list(row) for row in detalle]
            # Buscar la fila de horas ordinarias en el detalle
            for i, row in enumerate(detalle_modificable):
                if "ordinarias" in row[0].lower():
                    # Recalcular el subtotal de horas ordinarias como la mitad del salario
                    detalle_modificable[i][5] = salario_base / 2  # Actualizar el subtotal
                    break
            # Usar el detalle modificable para el resto del proceso
            detalle = detalle_modificable
            # Recalcular el total sumando todos los subtotales del detalle corregido
            total = sum(row[5] for row in detalle)
        else:
            total = resumen["total"]
        
        auxilio = resumen["auxilio"]
        salud = resumen["salud"]
        pension = resumen["pension"]
        deuda = deuda_valor if deuda_valor else 0
        neto = resumen["neto"]
        total_con_auxilio = total + auxilio

        for row in detalle:
            concept = row[0]
    
    # Ocultar "Horas extras diurnas" solo si es un fijo
            if tipo == "FIJO" and concept.strip().lower() == "horas extras diurnas":
                continue

            valor_hora = row[1]
            horas = row[2]
            valor_recargo = row[3]
            valor_total = row[4]
            subtotal = row[5]
            
            table_data.append([
                concept.upper(),
                f"${valor_hora:,.0f}",
                f"${valor_recargo:,.0f}",
                f"${valor_total:,.0f}",
                f"{horas:.1f}",
                f"${subtotal:,.0f}"
            ])


        row_heights = [20] + [20] * (len(table_data) - 1)
        table = Table(table_data, colWidths=[120, 70, 70, 70, 50, 80], rowHeights=row_heights)
        table.setStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('LEADING', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (5, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
            ('ALIGN', (1, 0), (3, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ])
        elements.append(table)
        elements.append(Spacer(1, 12))

        if auxilio > 0:
            elements.append(Paragraph(f"AUXILIO DE TRANSPORTE: ${auxilio:,.0f}", custom_style))
        if tipo == "FIJO":
            elements.append(Paragraph(f"APORTE SALUD: (${salud:,.0f})", custom_style))
            elements.append(Paragraph(f"APORTE PENSIÓN: (${pension:,.0f})", custom_style))
        if deuda_valor > 0:
            elements.append(Paragraph(f"MOTIVO DE DEUDA: {deuda_comentario.upper()}", custom_style))
            elements.append(Paragraph(f"DEUDA: (${deuda_valor:,.0f})", custom_style))
        elements.append(Paragraph(f"TOTAL (CON AUXILIO): ${total_con_auxilio:,.0f}", custom_style))
        neto_ajustado = max(0, neto)
        elements.append(Paragraph(f"TOTAL NETO A PAGAR: ${neto_ajustado:,.0f}", custom_style))
        elements.append(Spacer(1, 24))

        elements.append(Paragraph("FIRMA DEL TRABAJADOR: __________________________", custom_style))
        elements.append(Paragraph(f"CÉDULA: __________________________", custom_style))
        elements.append(Spacer(1, 12))

        doc.build(elements)
        registrar_log(f"PDF generado para {nombre} ({cedula})")
        messagebox.showinfo("ÉXITO", f"Comprobante guardado como:\n{archivo}")
    except Exception as e:
        registrar_log(f"Error generando PDF para {nombre}: {str(e)} - Detalle: {str(e.args)}")
        messagebox.showerror("ERROR", f"No se pudo generar el PDF: {str(e)} - Revisa el log {LOG_PATH}")

# Generar resúmenes interactivos HTML
def generar_resumen_html(tipo_empleado):
    empleados = obtener_empleados()
    if not empleados:
        messagebox.showwarning("Atención", "No hay empleados registrados.")
        return
    quincena = datetime.now().strftime("%Y-%m")
    empleados_filtrados = [emp for emp in empleados if emp[3] == tipo_empleado]
    if not empleados_filtrados:
        messagebox.showwarning("Atención", f"No hay empleados {tipo_empleado.lower()} registrados.")
        return

    archivo = HTML_RESUMEN_FIJO if tipo_empleado == "FIJO" else HTML_RESUMEN_TEMPORAL
    titulo = f"RESUMEN DE EMPLEADOS {tipo_empleado.upper()} - VILLA VENECIA"
    try:
        with open(archivo, "w", encoding="utf-8") as f:
            if tipo_empleado == "FIJO":
                f.write(f"""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{titulo}</title>
    <style>
        body {{ font-family: 'Arial', sans-serif; background-color: #F5F5F5; color: #333333; margin: 0; padding: 20px; text-align: center; }}
        .container {{ max-width: 900px; margin: auto; background: #FFFFFF; padding: 25px; border-radius: 10px; box-shadow: 0 0 10px rgba(0,0,0,0.1); overflow-x: auto; }}
        h1 {{ color: #26A69A; text-align: center; text-transform: uppercase; }}
        p {{ margin: 10px 0; text-align: center; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 14px; background-color: #FFFFFF; }}
        th, td {{ padding: 8px; text-align: center; border: 1px solid #E0E0E0; white-space: nowrap; }}
        th {{ background-color: #26A69A; color: #FFFFFF; font-weight: bold; }}
        tr:nth-child(even) {{ background-color: #F5F5F5; }}
        tr:nth-child(odd) {{ background-color: #FFFFFF; }}
        .total-row {{ background-color: #D3D3D3; font-weight: bold; }}
        .filter {{ margin-bottom: 20px; }}
    </style>
    <script>
        function filterEmployees() {{
            var filter = document.getElementById("typeFilter").value;
            var rows = document.getElementsByTagName("tr");
            for (var i = 1; i < rows.length - 1; i++) {{
                var type = rows[i].cells[0].innerText.split(" - ")[2];
                if (filter === "todos" || type === filter) {{
                    rows[i].style.display = "";
                }} else {{
                    rows[i].style.display = "none";
                }}
            }}
        }}
    </script>
</head>
<body>
    <div class="container">
        <h1>{titulo}</h1>
        <p>FECHA: {datetime.now().strftime('%m/%d/%Y %I:%M %p').upper()}</p>
        <div class="filter">
            <label for="typeFilter">Filtrar por tipo: </label>
            <select id="typeFilter" onchange="filterEmployees()">
                <option value="todos">Todos</option>
                <option value="{tipo_empleado}">Only {tipo_empleado}</option>
            </select>
        </div>
        <table>
            <tr>
                <th>Nombre - Cédula - Tipo</th>
                <th>Salario básico</th>
                <th>Días liquidados</th>
                <th>Salario devengado</th>
                <th>Horas extras Diurnas</th>
                <th>Recargos nocturnos Dominicales</th>
                <th>Extra nocturna</th>
                <th>Recargo nocturno</th>
                <th>Recargo Dominical laborado</th>
                <th>Hora diurna dominical o festiva</th>
                <th>Hora nocturna dominical o festiva</th>
                <th>Extra diurna dominical y festivo</th>
                <th>Extra nocturna dominical y festivo</th>
                <th>Auxilio de transporte</th>
                <th>Total devengado</th>
                <th>Salud</th>
                <th>Pensión</th>
                <th>Otras deducciones</th>
                <th>Neto pagado</th>
            </tr>
                """)
            else:
                f.write(f"""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{titulo}</title>
    <style>
        body {{ font-family: 'Arial', sans-serif; background-color: #F5F5F5; color: #333333; margin: 0; padding: 20px; text-align: center; }}
        .container {{ max-width: 900px; margin: auto; background: #FFFFFF; padding: 25px; border-radius: 10px; box-shadow: 0 0 10px rgba(0,0,0,0.1); overflow-x: auto; }}
        h1 {{ color: #26A69A; text-align: center; text-transform: uppercase; }}
        p {{ margin: 10px 0; text-align: center; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 14px; background-color: #FFFFFF; }}
        th, td {{ padding: 8px; text-align: center; border: 1px solid #E0E0E0; white-space: nowrap; }}
        th {{ background-color: #26A69A; color: #FFFFFF; font-weight: bold; }}
        tr:nth-child(even) {{ background-color: #F5F5F5; }}
        tr:nth-child(odd) {{ background-color: #FFFFFF; }}
        .total-row {{ background-color: #D3D3D3; font-weight: bold; }}
        .filter {{ margin-bottom: 20px; }}
    </style>
    <script>
        function filterEmployees() {{
            var filter = document.getElementById("typeFilter").value;
            var rows = document.getElementsByTagName("tr");
            for (var i = 1; i < rows.length - 1; i++) {{
                var type = rows[i].cells[0].innerText.split(" - ")[2];
                if (filter === "todos" || type === filter) {{
                    rows[i].style.display = "";
                }} else {{
                    rows[i].style.display = "none";
                }}
            }}
        }}
    </script>
</head>
<body>
    <div class="container">
        <h1>{titulo}</h1>
        <p>FECHA: {datetime.now().strftime('%m/%d/%Y %I:%M %p').upper()}</p>
        <div class="filter">
            <label for="typeFilter">Filtrar por tipo: </label>
            <select id="typeFilter" onchange="filterEmployees()">
                <option value="todos">Todos</option>
                <option value="{tipo_empleado}">Only {tipo_empleado}</option>
            </select>
        </div>
        <table>
            <tr>
                <th>Nombre - Cédula - Tipo</th>
                <th>Salario básico</th>
                <th>Salario devengado</th>
                <th>Horas extras Diurnas</th>
                <th>Recargos nocturnos Dominicales</th>
                <th>Extra nocturna</th>
                <th>Recargo nocturno</th>
                <th>Recargo Dominical laborado</th>
                <th>Hora diurna dominical o festivo</th>
                <th>Hora nocturna dominical o festivo</th>
                <th>Extra diurna dominical y festivo</th>
                <th>Extra nocturna dominical y festivo</th>
                <th>Otras deducciones</th>
                <th>Neto pagado</th>
            </tr>
                """)
            total_nomina = 0
            config = cargar_configuracion()
            for emp in empleados_filtrados:
                salario = emp[4]
                tipo = emp[3]
                horas_dict = obtener_horas_trabajadas(emp[0], quincena)
                detalle, resumen = calcular_valores(tipo, salario, horas_dict, config)
                valor_hora_base = salario / 220
                
                # Para empleados fijos, corregir el total de horas ordinarias
                if tipo == "FIJO":
                    # Convertir tuplas a listas para poder modificarlas
                    detalle_modificable = [list(row) for row in detalle]
                    # Buscar la fila de horas ordinarias en el detalle
                    for i, row in enumerate(detalle_modificable):
                        if "ordinarias" in row[0].lower():
                            # Recalcular el subtotal de horas ordinarias como la mitad del salario
                            detalle_modificable[i][5] = salario / 2  # Actualizar el subtotal
                            break
                    # Usar el detalle modificable para el resto del proceso
                    detalle = detalle_modificable
                    # Recalcular el total sumando todos los subtotales del detalle corregido
                    total = sum(row[5] for row in detalle)
                else:
                    total = resumen["total"]
                
                neto = resumen["neto"]
                deuda = float(horas_dict.get("deuda_valor", 0)) if horas_dict.get("deuda_valor") else 0
                valor_ordinarias = salario / 2 if tipo == "FIJO" else horas_dict.get("ordinarias", 0) * valor_hora_base
                valor_extras_diurnas = next((row[5] for row in detalle if row[0] == "Hora extra diurna"), 0)
                valor_recargos_nocturnos_dominicales = next((row[5] for row in detalle if row[0] == "Recargo nocturno dominical"), 0)
                valor_extra_nocturna = next((row[5] for row in detalle if row[0] == "Hora extra nocturna"), 0)
                valor_recargo_nocturno = next((row[5] for row in detalle if row[0] == "Recargo nocturno"), 0)
                valor_recargo_dominical = next((row[5] for row in detalle if row[0] == "Recargo diurno dominical"), 0)
                valor_hora_diurna_dominical = next((row[5] for row in detalle if row[0] == "Hora diurna dominical o festivo"), 0)
                valor_hora_nocturna_dominical = next((row[5] for row in detalle if row[0] == "Hora nocturna dominical o festivo"), 0)
                valor_extra_diurna_dominical = next((row[5] for row in detalle if row[0] == "Hora extra diurna dominical o festivo"), 0)
                valor_extra_nocturna_dominical = next((row[5] for row in detalle if row[0] == "Hora extra nocturna dominical o festivo"), 0)
                auxilio = config["auxilio_transporte"] if tipo == "FIJO" and salario <= 2 * config["salario_minimo"] and horas_dict.get("ordinarias", 0) > 0 else 0
                total_con_auxilio = total + auxilio
                dias_liquidados = 15 if tipo == "FIJO" else 0
                neto_pagado = total_con_auxilio - resumen["salud"] - resumen["pension"] - deuda
                if tipo_empleado == "FIJO":
                    f.write(f"<tr><td style='text-align:left;'>{emp[1].upper()} - {emp[2]} - {tipo.upper()}</td><td>${salario:,.0f}</td><td>{dias_liquidados}</td><td>${valor_ordinarias:,.0f}</td><td>${valor_extras_diurnas:,.0f}</td><td>${valor_recargos_nocturnos_dominicales:,.0f}</td><td>${valor_extra_nocturna:,.0f}</td><td>${valor_recargo_nocturno:,.0f}</td><td>${valor_recargo_dominical:,.0f}</td><td>${valor_hora_diurna_dominical:,.0f}</td><td>${valor_hora_nocturna_dominical:,.0f}</td><td>${valor_extra_diurna_dominical:,.0f}</td><td>${valor_extra_nocturna_dominical:,.0f}</td><td>${auxilio:,.0f}</td><td>${total_con_auxilio:,.0f}</td><td>${resumen['salud']:,.0f}</td><td>${resumen['pension']:,.0f}</td><td>${deuda:,.0f}</td><td>${neto_pagado:,.0f}</td></tr>\n")
                else:
                    f.write(f"<tr><td style='text-align:left;'>{emp[1].upper()} - {emp[2]} - {tipo.upper()}</td><td>${salario:,.0f}</td><td>${valor_ordinarias:,.0f}</td><td>${valor_extras_diurnas:,.0f}</td><td>${valor_recargos_nocturnos_dominicales:,.0f}</td><td>${valor_extra_nocturna:,.0f}</td><td>${valor_recargo_nocturno:,.0f}</td><td>${valor_recargo_dominical:,.0f}</td><td>${valor_hora_diurna_dominical:,.0f}</td><td>${valor_hora_nocturna_dominical:,.0f}</td><td>${valor_extra_diurna_dominical:,.0f}</td><td>${valor_extra_nocturna_dominical:,.0f}</td><td>${deuda:,.0f}</td><td>${neto:,.0f}</td></tr>\n")
                total_nomina += neto_pagado
            # Calcular totales basados en los datos generados
            total_row_data = []
            for emp in empleados_filtrados:
                salario = emp[4]
                tipo = emp[3]
                horas_dict = obtener_horas_trabajadas(emp[0], quincena)
                detalle, resumen = calcular_valores(tipo, salario, horas_dict, config)
                valor_hora_base = salario / 220
                
                # Para empleados fijos, corregir el total de horas ordinarias
                if tipo == "FIJO":
                    # Convertir tuplas a listas para poder modificarlas
                    detalle_modificable = [list(row) for row in detalle]
                    # Buscar la fila de horas ordinarias en el detalle
                    for i, row in enumerate(detalle_modificable):
                        if "ordinarias" in row[0].lower():
                            # Recalcular el subtotal de horas ordinarias como la mitad del salario
                            detalle_modificable[i][5] = salario / 2  # Actualizar el subtotal
                            break
                    # Usar el detalle modificable para el resto del proceso
                    detalle = detalle_modificable
                    # Recalcular el total sumando todos los subtotales del detalle corregido
                    total = sum(row[5] for row in detalle)
                else:
                    total = resumen["total"]
                
                neto = resumen["neto"]
                deuda = float(horas_dict.get("deuda_valor", 0)) if horas_dict.get("deuda_valor") else 0
                valor_ordinarias = salario / 2 if tipo == "FIJO" else horas_dict.get("ordinarias", 0) * valor_hora_base
                valor_extras_diurnas = next((row[5] for row in detalle if row[0] == "Hora extra diurna"), 0)
                valor_recargos_nocturnos_dominicales = next((row[5] for row in detalle if row[0] == "Recargo nocturno dominical"), 0)
                valor_extra_nocturna = next((row[5] for row in detalle if row[0] == "Hora extra nocturna"), 0)
                valor_recargo_nocturno = next((row[5] for row in detalle if row[0] == "Recargo nocturno"), 0)
                valor_recargo_dominical = next((row[5] for row in detalle if row[0] == "Recargo diurno dominical"), 0)
                valor_hora_diurna_dominical = next((row[5] for row in detalle if row[0] == "Hora diurna dominical o festivo"), 0)
                valor_hora_nocturna_dominical = next((row[5] for row in detalle if row[0] == "Hora nocturna dominical o festivo"), 0)
                valor_extra_diurna_dominical = next((row[5] for row in detalle if row[0] == "Hora extra diurna dominical o festivo"), 0)
                valor_extra_nocturna_dominical = next((row[5] for row in detalle if row[0] == "Hora extra nocturna dominical o festivo"), 0)
                auxilio = config["auxilio_transporte"] if tipo == "FIJO" and salario <= 2 * config["salario_minimo"] and horas_dict.get("ordinarias", 0) > 0 else 0
                total_con_auxilio = total + auxilio
                dias_liquidados = 15 if tipo == "FIJO" else 0
                neto_pagado = total_con_auxilio - resumen["salud"] - resumen["pension"] - deuda
                total_row_data.append([salario, dias_liquidados, valor_ordinarias, valor_extras_diurnas, valor_recargos_nocturnos_dominicales, valor_extra_nocturna, valor_recargo_nocturno, valor_recargo_dominical, valor_hora_diurna_dominical, valor_hora_nocturna_dominical, valor_extra_diurna_dominical, valor_extra_nocturna_dominical, auxilio, total_con_auxilio, resumen["salud"], resumen["pension"], deuda, neto_pagado])

            f.write(f"""
                <tr class="total-row"><td>TOTAL {tipo_empleado.upper()}</td><td>${sum(float(row[0]) for row in total_row_data):,.0f}</td><td>{sum(int(row[1]) for row in total_row_data)}</td><td>${sum(float(row[2]) for row in total_row_data):,.0f}</td><td>${sum(float(row[3]) for row in total_row_data):,.0f}</td><td>${sum(float(row[4]) for row in total_row_data):,.0f}</td><td>${sum(float(row[5]) for row in total_row_data):,.0f}</td><td>${sum(float(row[6]) for row in total_row_data):,.0f}</td><td>${sum(float(row[7]) for row in total_row_data):,.0f}</td><td>${sum(float(row[8]) for row in total_row_data):,.0f}</td><td>${sum(float(row[9]) for row in total_row_data):,.0f}</td><td>${sum(float(row[10]) for row in total_row_data):,.0f}</td><td>${sum(float(row[11]) for row in total_row_data):,.0f}</td><td>${sum(float(row[12]) for row in total_row_data):,.0f}</td><td>${sum(float(row[13]) for row in total_row_data):,.0f}</td><td>${sum(float(row[14]) for row in total_row_data):,.0f}</td><td>${sum(float(row[15]) for row in total_row_data):,.0f}</td><td>${sum(float(row[16]) for row in total_row_data):,.0f}</td><td>${sum(float(row[17]) for row in total_row_data):,.0f}</td></tr>
            </table>
        </div>
    </body>
    </html>
            """)
        messagebox.showinfo("ÉXITO", f"Resumen interactivo guardado en {archivo}")
        registrar_log(f"Resumen interactivo exportado a {archivo}")
        webbrowser.open(f"file://{os.path.abspath(archivo)}")
    except Exception as e:
        registrar_log(f"Error generando resumen HTML para {tipo_empleado}: {str(e)}")
        messagebox.showerror("ERROR", f"No se pudo generar el resumen interactivo: {str(e)}")

# Interfaz principal
class PayrollApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Nómina 2025 - Villa Venecia")
        self.root.geometry("1200x800")
        self.root.configure(bg="#FFFFFF")
        self.root.state('zoomed')

        try:
            self.root.iconbitmap(resource_path("nomina.ico"))
        except Exception as e:
            registrar_log(f"Error cargando ícono: {str(e)}")
            messagebox.showwarning("Advertencia", "Ícono nomina.ico no encontrado")

        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.style.configure("TButton", background="#26A69A", foreground="#FFFFFF", font=("Helvetica", 12, "bold"), padding=10)
        self.style.map("TButton", background=[('active', '#00796B')])
        self.style.configure("TLabel", background="#FFFFFF", font=("Helvetica", 12), foreground="#26A69A")
        self.style.configure("TFrame", background="#FFFFFF")
        self.style.configure("TCombobox", fieldbackground="#FFFFFF", foreground="#26A69A", background="#FFFFFF", font=("Helvetica", 12))
        self.style.configure("Header.TLabel", font=("Helvetica", 18, "bold"), foreground="#26A69A")
        self.style.configure("Treeview", font=("Helvetica", 10))
        self.style.configure("Treeview.Heading", font=("Helvetica", 10, "bold"))

        main_frame = ttk.Frame(root, padding="40")
        main_frame.pack(fill="both", expand=True)

        self.logo = tk.Label(main_frame, text="VILLA VENECIA", fg="#26A69A", bg="#FFFFFF", font=("Helvetica", 24, "bold"))
        self.logo.pack(pady=(0, 30))
        self.animation_index = 0
        self.animations_set_1 = [
            (lambda: self.animate_color_shift("🏛️ BIENVENIDO", ["#26A69A", "#00796B", "#004D40", "#26A69A"], 24), 1600),
            (lambda: self.animate_typing("💼 VILLA VENECIA", "#26A69A", 24), 3000),
            (lambda: self.animate_blink("📊 SISTEMA DE NÓMINA", "#00796B", 22), 2000),
            (lambda: self.animate_word_cycle(["💎 NÓMINA", "📅 HORARIOS", "💳 PAGOS", "👥 EMPLEADOS"], "#26A69A", 24), 2000),
        ]
        self.animations_set_2 = [
            (lambda: self.animate_logo("V. VENECIA", "#26A69A", 20), 2000),
            (lambda: self.animate_logo("VILLA VENECIA 🌟", "#26A69A", 26), 2500),
            (lambda: self.animate_logo(f"{random.choice(['💼', '🏢'])} V. V.", random.choice(["#26A69A", "#00796B", "#004D40"]), random.randint(20, 26)), 2000),
        ]
        self.current_animation_list = self.animations_set_1
        self.update_animation()

        self.clock = ttk.Label(main_frame, text="", font=("Helvetica", 12), foreground="#26A69A", background="#FFFFFF")
        self.clock.pack(pady=5)
        self.update_clock()

        self.stats_frame = ttk.Frame(main_frame, padding="10", relief="flat", borderwidth=1)
        self.stats_frame.pack(fill="x", pady=10)
        self.total_nomina_fijos_var = tk.StringVar(value="$0")
        self.total_nomina_temporales_var = tk.StringVar(value="$0")
        self.total_nomina_general_var = tk.StringVar(value="$0")
        self.total_empleados_var = tk.StringVar(value="0")
        self.actualizar_estadísticas()

        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.pack(fill="x", pady=20)
        buttons = [
            ("➕ Registrar Empleado", self.registrar_ui),
            ("🕒 Gestionar Horas Trabajadas", self.horas_ui),
            ("📁 Abrir Comprobantes PDF", self.ver_pdfs),
            ("📋 Exportar Resumen", self.exportar_resumen),
            ("🌐 Resumen Fijos", lambda: generar_resumen_html("FIJO")),
            ("🌐 Resumen Temporales", lambda: generar_resumen_html("TEMPORAL")),
            ("⚙ Actualizar Configuración", self.actualizar_config),
            ("🔄 Reiniciar Total Nómina", self.reiniciar_total_nomina),
            ("🔄 Reiniciar Resumen", self.reiniciar_total_nomina),
            ("🔁 Actualizar Nómina Fijos", self.actualizar_nomina_fijos_ui),
            ("🔁 Actualizar Nómina Temporales", self.actualizar_nomina_temporales_ui),
        ]
        for i, (text, command) in enumerate(buttons):
            btn = ttk.Button(buttons_frame, text=text, command=lambda cmd=command: self._execute_command(cmd), style="TButton", state="disabled" if command is None else "normal")
            btn.pack(fill="x", pady=5, padx=5)

    # Métodos existentes (sin cambios)
    def _execute_command(self, command):
        try:
            if command:
                command()
        except Exception as e:
            registrar_log(f"Error ejecutando comando: {str(e)}")
            messagebox.showerror("ERROR", f"Se produjo un error: {str(e)} - Revisa el log {LOG_PATH}")

    def animate_logo(self, text, color, size, blink=False):
        self.logo.config(text=text, fg=color, font=("Helvetica", size, "bold"))
        if blink:
            self.root.after(500, lambda: self.logo.config(fg="#B2DFDB"))
            self.root.after(1000, lambda: self.logo.config(fg=color))

    def update_animation(self):
        if self.animation_index >= len(self.current_animation_list):
            self.animation_index = 0
            self.current_animation_list = (
                self.animations_set_2 if self.current_animation_list == self.animations_set_1 else self.animations_set_1
            )
        animation_func, delay = self.current_animation_list[self.animation_index]
        animation_func()
        self.animation_index += 1
        self.root.after(delay, self.update_animation)

    def animate_typing(self, full_text, color, size=24, index=0):
        if index <= len(full_text):
            self.logo.configure(text=full_text[:index], foreground=color, font=("Helvetica", size, "bold"))
            self.root.after(100, lambda: self.animate_typing(full_text, color, size, index + 1))

    def animate_blink(self, text, color, size=24, count=0):
        if count < 6:
            new_color = color if count % 2 == 0 else "#FFFFFF"
            self.logo.configure(text=text, foreground=new_color, font=("Helvetica", size, "bold"))
            self.root.after(300, lambda: self.animate_blink(text, color, size, count + 1))

    def animate_pulse(self, text, color, size=24, grow=True, step=0):
        new_size = size + step if grow else size - step
        self.logo.configure(text=text, foreground=color, font=("Helvetica", new_size, "bold"))
        if step < 4:
            self.root.after(100, lambda: self.animate_pulse(text, color, size, grow, step + 1))
        else:
            if grow:
                self.animate_pulse(text, color, size, grow=False)

    def animate_color_shift(self, text, colors, index=0, size=24):
        if index < len(colors):
            self.logo.configure(text=text, foreground=colors[index], font=("Helvetica", size, "bold"))
            self.root.after(200, lambda: self.animate_color_shift(text, colors, index + 1))

    def animate_word_cycle(self, words, color="#26A69A", size=24, index=0):
        if index < len(words):
            self.logo.configure(text=words[index], foreground=color, font=("Helvetica", size, "bold"))
            self.root.after(400, lambda: self.animate_word_cycle(words, color, size, index + 1))

    def update_clock(self):
        dias_semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        dia = dias_semana[datetime.now().weekday()]
        current_time = datetime.now().strftime("%I:%M %p")
        self.clock.config(text=f"Hora: {current_time} - {dia}, {datetime.now().strftime('%d/%m/%Y')}")
        self.root.after(1000, self.update_clock)

    def actualizar_estadísticas(self):
        for widget in self.stats_frame.winfo_children():
            widget.destroy()
        empleados = obtener_empleados()
        total_nomina_fijos = 0
        total_nomina_temporales = 0
        total_nomina_general = 0
        total_empleados = len(empleados)
        config = cargar_configuracion()
        quincena = datetime.now().strftime("%Y-%m")
        for emp in empleados:
            salario = emp[4]
            tipo = emp[3]
            horas_dict = obtener_horas_trabajadas(emp[0], quincena)
            if horas_dict.get("ordinarias", 0) > 0 or any(horas_dict.get(th, 0) > 0 for th, _ in TIPOS_HORAS[1:]):
                detalle, resumen = calcular_valores(tipo, salario, horas_dict, config)
                neto = resumen["neto"]
                if tipo == "FIJO":
                    total_nomina_fijos += neto
                else:
                    total_nomina_temporales += neto
                total_nomina_general += neto
        self.total_empleados_var.set(str(total_empleados))
        self.total_nomina_fijos_var.set(f"${max(0, total_nomina_fijos):,.0f}")
        self.total_nomina_temporales_var.set(f"${max(0, total_nomina_temporales):,.0f}")
        self.total_nomina_general_var.set(f"${max(0, total_nomina_general):,.0f}")
        ttk.Label(self.stats_frame, text=f"Empleados Registrados: {self.total_empleados_var.get()}", style="TLabel").pack(side="left", padx=10)
        ttk.Label(self.stats_frame, text=f"Total Nómina Fijos: {self.total_nomina_fijos_var.get()}", style="TLabel").pack(side="left", padx=10)
        ttk.Label(self.stats_frame, text=f"Total Nómina Temporales: {self.total_nomina_temporales_var.get()}", style="TLabel").pack(side="left", padx=10)
        ttk.Label(self.stats_frame, text=f"Total Nómina General: {self.total_nomina_general_var.get()}", style="TLabel").pack(side="left", padx=10)

    def reiniciar_total_nomina(self):
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("DELETE FROM horas_trabajadas WHERE quincena = ?", (datetime.now().strftime("%Y-%m"),))
        conn.commit()
        conn.close()
        self.actualizar_estadísticas()
        messagebox.showinfo("ÉXITO", "Totales de nómina reiniciados para la quincena actual")

    def registrar_ui(self):
        w = tk.Toplevel(self.root)
        w.title("Registrar Empleado")
        w.geometry("600x750")
        w.configure(bg="#FFFFFF")

        main_frame = ttk.Frame(w, padding="30")
        main_frame.pack(expand=True)
        main_frame.columnconfigure(1, weight=1)

        campos = ["NOMBRE:", "CÉDULA:", "TIPO:", "SALARIO BASE:", "COMENTARIO:"]
        entries = {}
        for i, t in enumerate(campos):
            label = ttk.Label(main_frame, text=t, style="TLabel")
            label.grid(row=i, column=0, sticky="e", padx=15, pady=20)
            entry = ttk.Entry(main_frame, width=45, font=("Helvetica", 12))
            if t == "TIPO:":
                entry = ttk.Combobox(main_frame, values=["FIJO", "TEMPORAL"], state="readonly", width=43, style="TCombobox")
                entry.current(0)
            entry.grid(row=i, column=1, padx=15, pady=20, sticky="ew")
            entries[t.replace(":", "")] = entry

        def guardar(event=None):
            nombre = entries["NOMBRE"].get().strip()
            cedula = entries["CÉDULA"].get().strip()
            tipo = entries["TIPO"].get().strip()
            salario = entries["SALARIO BASE"].get().strip()
            comentario = entries["COMENTARIO"].get().strip()

            if guardar_empleado(nombre, cedula, tipo, salario, comentario):
                messagebox.showinfo("ÉXITO", "Empleado registrado con éxito")
                self.actualizar_estadísticas()
                w.destroy()

        ttk.Button(main_frame, text="Guardar", command=guardar, style="TButton").grid(row=len(campos), column=0, columnspan=2, pady=30, sticky="nsew")
        w.bind('<Return>', guardar)

    def horas_ui(self):
        w = tk.Toplevel(self.root)
        w.title("Gestionar Horas Trabajadas")
        w.geometry("900x1000")
        w.configure(bg="#FFFFFF")

        main_frame = ttk.Frame(w, padding="30")
        main_frame.pack(fill="both", expand=True)

        empleados = obtener_empleados()
        if not empleados:
            messagebox.showwarning("Atención", "Primero debes registrar empleados.")
            w.destroy()
            return

        ttk.Label(main_frame, text="EMPLEADO:", style="TLabel").grid(row=0, column=0, sticky="e", padx=15, pady=20)
        emp_combo = ttk.Combobox(main_frame, values=[f"{e[1]} - {e[2]}" for e in empleados], state="readonly", width=50, style="TCombobox")
        emp_combo.grid(row=0, column=1, padx=15, pady=20, sticky="ew")

        entries_frame = ttk.Frame(main_frame)
        entries_frame.grid(row=1, column=0, columnspan=2, padx=15, pady=20, sticky="nsew")
        self.entries = {}

        def update_form(event):
            for widget in entries_frame.winfo_children():
                widget.destroy()

            if emp_combo.current() == -1:
                return

            idx = emp_combo.current()
            empleado_id = empleados[idx][0]
            quincena = datetime.now().strftime("%Y-%m")
            horas_dict = obtener_horas_trabajadas(empleado_id, quincena)
            self.entries.clear()

            ttk.Label(entries_frame, text="HORAS ORDINARIAS:", style="TLabel").grid(row=0, column=0, sticky="e", padx=15)
            self.entries["ordinarias"] = ttk.Entry(entries_frame, width=30, font=("Helvetica", 12))
            self.entries["ordinarias"].insert(0, str(horas_dict.get("ordinarias", 0)))
            self.entries["ordinarias"].grid(row=0, column=1, padx=15, pady=15)

            start_row = 1
            for i, (th, _) in enumerate(TIPOS_HORAS[1:], start_row):
                ttk.Label(entries_frame, text=th.replace("_", " ").capitalize(), style="TLabel").grid(row=i, column=0, sticky="e", padx=15)
                self.entries[th] = ttk.Entry(entries_frame, width=30, font=("Helvetica", 12))
                self.entries[th].insert(0, str(horas_dict.get(th, 0)))
                self.entries[th].grid(row=i, column=1, padx=15, pady=15)

            ttk.Label(entries_frame, text="MOTIVO DE DEUDA (OPCIONAL):", style="TLabel").grid(row=len(TIPOS_HORAS) + 1, column=0, sticky="e", padx=15)
            self.entries["deuda_comentario"] = ttk.Entry(entries_frame, width=30, font=("Helvetica", 12))
            self.entries["deuda_comentario"].insert(0, horas_dict.get("deuda_comentario", ""))
            self.entries["deuda_comentario"].grid(row=len(TIPOS_HORAS) + 1, column=1, padx=15, pady=15)
            ttk.Label(entries_frame, text="VALOR DE DEUDA (OPCIONAL):", style="TLabel").grid(row=len(TIPOS_HORAS) + 2, column=0, sticky="e", padx=15)
            self.entries["deuda_valor"] = ttk.Entry(entries_frame, width=30, font=("Helvetica", 12))
            self.entries["deuda_valor"].insert(0, str(horas_dict.get("deuda_valor", 0)))
            self.entries["deuda_valor"].grid(row=len(TIPOS_HORAS) + 2, column=1, padx=15, pady=15)

            ttk.Button(entries_frame, text="Limpiar Quincena", command=self.limpiar_quincena, style="TButton").grid(row=len(TIPOS_HORAS) + 3, column=0, columnspan=2, pady=15, sticky="nsew")

            entries_frame.columnconfigure(1, weight=1)

        emp_combo.bind("<<ComboboxSelected>>", update_form)

        def procesar(event=None):
            if emp_combo.current() == -1:
                messagebox.showwarning("Atención", "Selecciona un empleado")
                return

            idx = emp_combo.current()
            emp = empleados[idx]
            empleado_id, nombre, cedula, tipo, salario = emp[0], emp[1], emp[2], emp[3], emp[4]
            horas_dict = {}
            quincena = datetime.now().strftime("%Y-%m")

            try:
                for th in self.entries:
                    value = self.entries[th].get().strip()
                    if th in ["deuda_comentario"]:
                        horas_dict[th] = value
                    else:
                        horas_dict[th] = float(value) if value else 0
            except ValueError:
                messagebox.showerror("ERROR", "Las horas y el valor de deuda deben ser números válidos")
                return

            guardar_horas_trabajadas(empleado_id, horas_dict, quincena)
            deuda_comentario = horas_dict.get("deuda_comentario", "")
            deuda_valor = horas_dict.get("deuda_valor", 0)
            detalle, resumen = calcular_valores(tipo, salario, horas_dict, cargar_configuracion())
            horas_dict["total_horas"] = resumen["total_horas"]
            generar_pdf(nombre, cedula, tipo, salario, horas_dict, cargar_configuracion(), deuda_comentario, deuda_valor)
            self.actualizar_estadísticas()

        ttk.Button(main_frame, text="Generar PDF", command=procesar, style="TButton").grid(row=2, column=0, columnspan=2, pady=30, padx=15, sticky="nsew")
        w.bind('<Return>', procesar)
        main_frame.columnconfigure(1, weight=1)

    def limpiar_quincena(self):
        for th in self.entries:
            self.entries[th].delete(0, tk.END)
            if th not in ["deuda_comentario"]:
                self.entries[th].insert(0, "0")
        self.actualizar_estadísticas()

    def ver_pdfs(self):
        if not os.path.exists(PDF_DIR):
            messagebox.showinfo("Información", "Aún no hay PDFs generados.")
            return
        archivos = [f for f in os.listdir(PDF_DIR) if f.endswith(".pdf")]
        if not archivos:
            messagebox.showinfo("Información", "No se encontraron archivos PDF.")
            return
        w = tk.Toplevel(self.root)
        w.title("Comprobantes PDF - Villa Venecia")
        w.geometry("1000x750")
        w.configure(bg="#FFFFFF")

        main_frame = ttk.Frame(w, padding="30")
        main_frame.pack(fill="both", expand=True)

        ttk.Label(main_frame, text="Selecciona un comprobante para abrir", style="TLabel").pack(pady=20)
        lista = tk.Listbox(main_frame, width=80, bg="#FFFFFF", fg="#26A69A", font=("Helvetica", 12))
        for a in archivos:
            lista.insert(tk.END, a)
        lista.pack(pady=20, fill="both", expand=True)

        def abrir(event):
            sel = lista.curselection()
            if sel:
                archivo = archivos[sel[0]]
                ruta = os.path.abspath(os.path.join(PDF_DIR, archivo))
                webbrowser.open(f"file://{ruta}")

        lista.bind("<Double-Button-1>", abrir)
        ttk.Button(main_frame, text="Abrir Selección", command=lambda: abrir(None), style="TButton").pack(pady=30, side="bottom", anchor="e")

    def exportar_resumen(self):
        try:
            registrar_log("Intentando exportar resumen...")
            empleados = obtener_empleados()
            if not empleados:
                messagebox.showwarning("Atención", "No hay empleados registrados.")
                return
            quincena = datetime.now().strftime("%Y-%m")
            data = []
            total_nomina = 0
            config = cargar_configuracion()
            empleados_fijos = [emp for emp in empleados if emp[3] == "FIJO"]
            if not empleados_fijos:
                messagebox.showwarning("Atención", "No hay empleados fijos registrados.")
                return

            # Encabezados (solo una fila)
            columns = ["Nombre del empleado", "Cedula del empleado", "Salario básico", "Días liquidados", "Salario devengado",
                       "Horas extras Diurnas", "Recargos nocturnos Dominicales", "Extra nocturna", "Recargo nocturno",
                       "Recargo Dominical laborado", "Hora diurna dominical o festiva", "Hora nocturna dominical o festiva",
                       "Extra diurna dominical y festivo", "Extra nocturna dominical y festivo", "Auxilio de transporte",
                       "Total devengado", "Salud", "Pensión", "Otras deducciones", "Neto pagado", "Firma del empleado"]
            data.append(columns)

            # Datos específicos
            for emp in empleados_fijos:
                salario = emp[4]
                tipo = emp[3]
                horas_dict = obtener_horas_trabajadas(emp[0], quincena)
                detalle, resumen = calcular_valores(tipo, salario, horas_dict, config)
                valor_hora_base = salario / 220
                total = resumen["total"]
                neto = resumen["neto"]
                deuda = float(horas_dict.get("deuda_valor", 0)) if horas_dict.get("deuda_valor") else 0
                valor_ordinarias = salario / 2 if tipo == "FIJO" else horas_dict.get("ordinarias", 0) * valor_hora_base
                valor_extras_diurnas = next((row[5] for row in detalle if row[0] == "Hora extra diurna"), 0)
                valor_recargos_nocturnos_dominicales = next((row[5] for row in detalle if row[0] == "Recargo nocturno dominical"), 0)
                valor_extra_nocturna = next((row[5] for row in detalle if row[0] == "Hora extra nocturna"), 0)
                valor_recargo_nocturno = next((row[5] for row in detalle if row[0] == "Recargo nocturno"), 0)
                valor_recargo_dominical = next((row[5] for row in detalle if row[0] == "Recargo diurno dominical"), 0)
                valor_hora_diurna_dominical = next((row[5] for row in detalle if row[0] == "Hora diurna dominical o festivo"), 0)
                valor_hora_nocturna_dominical = next((row[5] for row in detalle if row[0] == "Hora nocturna dominical o festivo"), 0)
                valor_extra_diurna_dominical = next((row[5] for row in detalle if row[0] == "Hora extra diurna dominical o festivo"), 0)
                valor_extra_nocturna_dominical = next((row[5] for row in detalle if row[0] == "Hora extra nocturna dominical o festivo"), 0)
                auxilio = config["auxilio_transporte"] if tipo == "FIJO" and salario <= 2 * config["salario_minimo"] and horas_dict.get("ordinarias", 0) > 0 else 0
                total_con_auxilio = total + auxilio
                dias_liquidados = 15 if tipo == "FIJO" else 0
                neto_pagado = total_con_auxilio - resumen["salud"] - resumen["pension"] - deuda
                data.append([
                    emp[1].upper(), emp[2], salario, dias_liquidados, valor_ordinarias, valor_extras_diurnas,
                    valor_recargos_nocturnos_dominicales, valor_extra_nocturna, valor_recargo_nocturno,
                    valor_recargo_dominical, valor_hora_diurna_dominical, valor_hora_nocturna_dominical,
                    valor_extra_diurna_dominical, valor_extra_nocturna_dominical, auxilio, total_con_auxilio,
                    resumen["salud"], resumen["pension"], deuda, neto_pagado, ""
                ])
                total_nomina += neto_pagado

            # Agregar filas vacías
            for _ in range(4):  # row3 a row6
                data.append([""] * 21)

            # Fila de totales con autosuma
            total_row = ["Total General"]
            for i in range(1, len(columns)):
                if i in [2, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19]:  # Columnas numéricas
                    total = sum(float(str(row[i]).replace('$', '').replace(',', '')) if row[i] and isinstance(row[i], str) and row[i].replace('.', '').replace('-', '').replace('$', '').replace(',', '').replace(' ', '').isdigit() else float(row[i]) if row[i] else 0 for row in data[1:-1])
                    total_row.append(f"${total:,.2f}" if i != 3 else f"{int(total)}")
                else:
                    total_row.append("")
            data.append(total_row)

            # Crear DataFrame y guardar en Excel con formato
            df = pd.DataFrame(data, columns=columns)
            df.to_excel("resumen_empleados_fijos.xlsx", index=False, sheet_name="NÓMINA", float_format="%.2f")

            # Aplicar formato al Excel
            wb = load_workbook("resumen_empleados_fijos.xlsx")
            ws = wb["NÓMINA"]

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for row in ws.rows:
                for cell in row:
                    cell.border = thin_border
                    if cell.column > 1:  # Excluir la columna A (texto)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        if cell.row > 1 and cell.column_letter != "D":  # Excluir "Días liquidados"
                            cell.number_format = "$#,##0.00"

            for cell in ws[1]:  # Fila 1 contiene los encabezados
                cell.font = Font(bold=True)
            for cell in ws[ws.max_row]:  # Fila de totales
                cell.font = Font(bold=True)

            header_fill = PatternFill(start_color="003087", end_color="003087", fill_type="solid")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF")

            for i, row in enumerate(ws.rows):
                if i > 0 and i < ws.max_row - 1:
                    fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid") if i % 2 == 0 else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    for cell in row:
                        cell.fill = fill
            total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            for cell in ws[ws.max_row]:
                cell.fill = total_fill

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save("resumen_empleados_fijos.xlsx")

            # Exportar resumen tradicional a texto
            with open(RESUMEN_PATH, "w", encoding="utf-8") as f:
                f.write(f"RESUMEN DE EMPLEADOS - {datetime.now().strftime('%m/%d/%Y %I:%M %p').upper()}\n")
                f.write("=" * 60 + "\n")
                for emp in empleados:
                    salario = emp[4]
                    tipo = emp[3]
                    horas_dict = obtener_horas_trabajadas(emp[0], quincena)
                    detalle, resumen = calcular_valores(tipo, salario, horas_dict, config)
                    total_horas = resumen["total_horas"]
                    neto = resumen["neto"]
                    if tipo == "FIJO" and salario <= 2 * config["salario_minimo"]:
                        neto += config["auxilio_transporte"]
                    f.write(f"NOMBRE: {emp[1].upper()}\nCÉDULA: {emp[2].upper()}\nTIPO: {tipo.upper()}\nSALARIO: ${salario:,.0f}\nHORAS: {total_horas:.1f}\nNETO: ${neto:,.0f}\nCOMENTARIO: {emp[5].upper() or 'N/A'}\n\n")
            messagebox.showinfo("ÉXITO", f"Resumen guardado en resumen_empleados_fijos.xlsx y {RESUMEN_PATH}")
            registrar_log(f"Resumen exportado a Excel y texto")
        except Exception as e:
            registrar_log(f"Error exportando resumen: {str(e)}")
            messagebox.showerror("ERROR", f"No se pudo exportar el resumen: {str(e)} - Revisa el log {LOG_PATH}")

    def actualizar_config(self):
        w = tk.Toplevel(self.root)
        w.title("Actualizar Configuración")
        w.geometry("600x750")
        w.configure(bg="#FFFFFF")

        main_frame = ttk.Frame(w, padding="30")
        main_frame.pack(expand=True)
        main_frame.columnconfigure(1, weight=1)

        campos = ["SALARIO MÍNIMO:", "AUXILIO DE TRANSPORTE:", "DESCUENTO SALUD (%):", "DESCUENTO PENSIÓN (%):", "NOMBRE EMPRESA:", "NIT:", "DIRECCIÓN:"]
        entries = {}
        config = cargar_configuracion()
        for i, t in enumerate(campos):
            label = ttk.Label(main_frame, text=t, style="TLabel")
            label.grid(row=i, column=0, sticky="e", padx=15, pady=20)
            entry = ttk.Entry(main_frame, width=45, font=("Helvetica", 12))
            if t == "SALARIO MÍNIMO:":
                entry.insert(0, str(config["salario_minimo"]))
            elif t == "AUXILIO DE TRANSPORTE:":
                entry.insert(0, str(config["auxilio_transporte"]))
            elif t == "DESCUENTO SALUD (%):":
                entry.insert(0, str(config["descuentos"]["salud"] * 100))
            elif t == "DESCUENTO PENSIÓN (%):":
                entry.insert(0, str(config["descuentos"]["pension"] * 100))
            elif t == "NOMBRE EMPRESA:":
                entry.insert(0, config["empresa"]["nombre"])
            elif t == "NIT:":
                entry.insert(0, config["empresa"]["nit"])
            elif t == "DIRECCIÓN:":
                entry.insert(0, config["empresa"]["direccion"])
            entry.grid(row=i, column=1, padx=15, pady=20, sticky="ew")
            entries[t.replace(":", "")] = entry

        def guardar_config(event=None):
            try:
                new_config = config.copy()
                salario_minimo = entries["SALARIO MÍNIMO"].get().strip()
                auxilio_transporte = entries["AUXILIO DE TRANSPORTE"].get().strip()
                descuento_salud = entries["DESCUENTO SALUD"].get().strip()
                descuento_pension = entries["DESCUENTO PENSIÓN"].get().strip()
                
                if salario_minimo:
                    salario_minimo = float(salario_minimo.replace(',', ''))
                    if salario_minimo <= 0:
                        raise ValueError("El salario mínimo debe ser mayor a 0")
                    new_config["salario_minimo"] = salario_minimo
                if auxilio_transporte:
                    auxilio_transporte = float(auxilio_transporte.replace(',', ''))
                    if auxilio_transporte < 0:
                        raise ValueError("El auxilio de transporte no puede ser negativo")
                    new_config["auxilio_transporte"] = auxilio_transporte
                if descuento_salud:
                    descuento_salud = float(descuento_salud.replace(',', '')) / 100
                    if descuento_salud < 0 or descuento_salud > 1:
                        raise ValueError("El descuento de salud debe estar entre 0% y 100%")
                    new_config["descuentos"]["salud"] = descuento_salud
                if descuento_pension:
                    descuento_pension = float(descuento_pension.replace(',', '')) / 100
                    if descuento_pension < 0 or descuento_pension > 1:
                        raise ValueError("El descuento de pensión debe estar entre 0% y 100%")
                    new_config["descuentos"]["pension"] = descuento_pension
                
                new_config["empresa"]["nombre"] = entries["NOMBRE EMPRESA"].get().strip() or config["empresa"]["nombre"]
                new_config["empresa"]["nit"] = entries["NIT"].get().strip() or config["empresa"]["nit"]
                new_config["empresa"]["direccion"] = entries["DIRECCIÓN"].get().strip() or config["empresa"]["direccion"]
                
                guardar_configuracion(new_config)
                messagebox.showinfo("ÉXITO", "Configuración actualizada con éxito")
                self.actualizar_estadísticas()
                w.destroy()
            except ValueError as e:
                messagebox.showerror("ERROR", f"Valores numéricos inválidos: {str(e)}")
            except Exception as e:
                messagebox.showerror("ERROR", f"Error al guardar configuración: {str(e)}")

        ttk.Button(main_frame, text="Guardar Cambios", command=guardar_config, style="TButton").grid(row=len(campos), column=0, columnspan=2, pady=30, sticky="nsew")
        w.bind('<Return>', guardar_config)

    # Nuevos métodos para actualizar las hojas de Excel
    def actualizar_nomina_fijos_ui(self):
        """Open UI to preview and update fixed employees' payroll Excel."""
        self._actualizar_nomina_ui("FIJO", "NOMINA-2025.xlsx", "NÓMINA")

    def actualizar_nomina_temporales_ui(self):
        """Open UI to preview and update temporary employees' payroll Excel."""
        self._actualizar_nomina_ui("TEMPORAL", "NOMINA-2025.xlsx", "NÓMINA TEMPORALES")

    def _actualizar_nomina_ui(self, tipo_empleado, archivo_excel, hoja):
        """Generic UI for previewing and updating payroll Excel for a given employee type."""
        w = tk.Toplevel(self.root)
        w.title(f"Actualizar Nómina {tipo_empleado} - Villa Venecia")
        w.geometry("1200x800")
        w.configure(bg="#FFFFFF")

        main_frame = ttk.Frame(w, padding="30")
        main_frame.pack(fill="both", expand=True)

        empleados = obtener_empleados()
        empleados_filtrados = [emp for emp in empleados if emp[3] == tipo_empleado]
        if not empleados_filtrados:
            messagebox.showwarning("Atención", f"No hay empleados {tipo_empleado.lower()} registrados.")
            w.destroy()
            return

        quincena = datetime.now().strftime("%Y-%m")
        config = cargar_configuracion()

        # Define columns based on employee type
        columns = [
            "Nombre del empleado", "Cedula del empleado", "Salario básico", "Días liquidados",
            "Salario devengado", "Horas extras Diurnas", "Recargos nocturnos Dominicales",
            "Extra nocturna", "Recargo nocturno", "Recargo Dominical laborado",
            "Hora diurna dominical o festiva", "Hora nocturna dominical o festiva",
            "Extra diurna dominical y festivo", "Extra nocturna dominical y festivo",
            "Auxilio de transporte", "Total devengado", "Menos salario oficial", "Salud",
            "Pensión", "Fondo de solidaridad pensional", "Retención en la fuente",
            "Otras deducciones", "Neto pagado", "Firma del empleado"
        ] if tipo_empleado == "FIJO" else [
            "Nombre del empleado", "Cedula del empleado", "Salario básico", "Salario devengado",
            "Horas extras Diurnas", "Recargos nocturnos Dominicales", "Extra nocturna",
            "Recargo nocturno", "Recargo Dominical laborado", "Hora diurna dominical o festiva",
            "Hora nocturna dominical o festiva", "Extra diurna dominical y festivo",
            "Extra nocturna dominical y festivo", "Auxilio de transporte", "Total devengado",
            "Menos salario oficial", "Salud", "Pensión", "Fondo de solidaridad pensional",
            "Retención en la fuente", "Otras deducciones", "Neto pagado", "Firma del empleado"
        ]

        # Collect data
        data = []
        total_nomina = 0
        for emp in empleados_filtrados:
            salario = emp[4]
            tipo = emp[3]
            horas_dict = obtener_horas_trabajadas(emp[0], quincena)
            detalle, resumen = calcular_valores(tipo, salario, horas_dict, config)
            valor_hora_base = salario / 220
            total = resumen["total"]
            neto = resumen["neto"]
            deuda = float(horas_dict.get("deuda_valor", 0)) if horas_dict.get("deuda_valor") else 0
            valor_ordinarias = salario / 2 if tipo == "FIJO" else horas_dict.get("ordinarias", 0) * valor_hora_base
            valor_extras_diurnas = next((row[5] for row in detalle if row[0] == "Hora extra diurna"), 0)
            valor_recargos_nocturnos_dominicales = next((row[5] for row in detalle if row[0] == "Recargo nocturno dominical"), 0)
            valor_extra_nocturna = next((row[5] for row in detalle if row[0] == "Hora extra nocturna"), 0)
            valor_recargo_nocturno = next((row[5] for row in detalle if row[0] == "Recargo nocturno"), 0)
            valor_recargo_dominical = next((row[5] for row in detalle if row[0] == "Recargo diurno dominical"), 0)
            valor_hora_diurna_dominical = next((row[5] for row in detalle if row[0] == "Hora diurna dominical o festivo"), 0)
            valor_hora_nocturna_dominical = next((row[5] for row in detalle if row[0] == "Hora nocturna dominical o festivo"), 0)
            valor_extra_diurna_dominical = next((row[5] for row in detalle if row[0] == "Hora extra diurna dominical o festivo"), 0)
            valor_extra_nocturna_dominical = next((row[5] for row in detalle if row[0] == "Hora extra nocturna dominical o festivo"), 0)
            auxilio = config["auxilio_transporte"] if tipo == "FIJO" and salario <= 2 * config["salario_minimo"] and horas_dict.get("ordinarias", 0) > 0 else 0
            total_con_auxilio = total + auxilio
            dias_liquidados = 15 if tipo == "FIJO" else 0
            menos_salario_oficial = 0  # No especificado en el Excel
            salud = resumen["salud"]
            pension = resumen["pension"]
            fondo_solidaridad = 0  # No especificado
            retencion_fuente = 0  # No especificado
            neto_pagado = total_con_auxilio - salud - pension - deuda - fondo_solidaridad - retencion_fuente
            row = [
                emp[1].upper(), emp[2], salario, dias_liquidados, valor_ordinarias,
                valor_extras_diurnas, valor_recargos_nocturnos_dominicales, valor_extra_nocturna,
                valor_recargo_nocturno, valor_recargo_dominical, valor_hora_diurna_dominical,
                valor_hora_nocturna_dominical, valor_extra_diurna_dominical, valor_extra_nocturna_dominical,
                auxilio, total_con_auxilio, menos_salario_oficial, salud, pension, fondo_solidaridad,
                retencion_fuente, deuda, neto_pagado, ""
            ] if tipo == "FIJO" else [
                emp[1].upper(), emp[2], salario, valor_ordinarias,
                valor_extras_diurnas, valor_recargos_nocturnos_dominicales, valor_extra_nocturna,
                valor_recargo_nocturno, valor_recargo_dominical, valor_hora_diurna_dominical,
                valor_hora_nocturna_dominical, valor_extra_diurna_dominical, valor_extra_nocturna_dominical,
                auxilio, total_con_auxilio, menos_salario_oficial, salud, pension, fondo_solidaridad,
                retencion_fuente, deuda, neto_pagado, ""
            ]
            data.append(row)
            total_nomina += neto_pagado

        # Add total row
        total_row = ["Totales"] + [""] * (len(columns) - 2)
        for i in range(2, len(columns) - 1):
            if i in [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21]:
                total = sum(float(row[i]) if row[i] else 0 for row in data)
                total_row[i] = total if i != 3 else int(total)
        total_row[-1] = ""
        data.append(total_row)

        # Create Treeview for preview
        ttk.Label(main_frame, text=f"Resumen Nómina {tipo_empleado}", style="Header.TLabel").pack(pady=20)
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=20)
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=100, anchor="center")
        tree.pack(side="left", fill="both", expand=True)

        # Add scrollbar
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        scrollbar.pack(side="right", fill="y")
        tree.configure(yscrollcommand=scrollbar.set)

        # Populate Treeview
        for row in data:
            tree.insert("", "end", values=row)

        def confirmar():
            try:
                wb = load_workbook(archivo_excel)
                ws = wb[hoja]
                # Limpiar filas a partir de la 8
                for row in ws["A8:W" + str(ws.max_row)]:
                    for cell in row:
                        cell.value = None
                # Escribir datos
                for i, row in enumerate(data, start=8):
                    for j, value in enumerate(row, start=1):
                        ws.cell(row=i, column=j).value = value
                # Aplicar formato
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                for row in ws["A6:W" + str(ws.max_row)]:
                    for cell in row:
                        cell.border = thin_border
                        if cell.column > 1 and cell.row > 6:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            if cell.column_letter not in ["A", "W"] and cell.row > 7:
                                cell.number_format = "$#,##0.00"
                # Formato de encabezados (fila 6)
                for cell in ws[6]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="003087", end_color="003087", fill_type="solid")
                    cell.font = Font(color="FFFFFF")
                # Formato de filas alternadas
                for i, row in enumerate(ws["A8:W" + str(ws.max_row)], start=8):
                    if i < ws.max_row:
                        fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid") if i % 2 == 0 else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                        for cell in row:
                            cell.fill = fill
                    else:
                        for cell in row:
                            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                            cell.font = Font(bold=True)
                # Ajustar ancho de columnas
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column_letter].width = adjusted_width
                wb.save(archivo_excel)
                messagebox.showinfo("ÉXITO", f"Nómina {tipo_empleado.lower()} actualizada en {archivo_excel}, hoja {hoja}")
                registrar_log(f"Nómina {tipo_empleado} actualizada en {archivo_excel}, hoja {hoja}")
                w.destroy()
            except Exception as e:
                registrar_log(f"Error actualizando nómina {tipo_empleado}: {str(e)}")
                messagebox.showerror("Error", f"No se pudo actualizar la nómina: {str(e)} - Revisa el log {LOG_PATH}")

        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=20)
        ttk.Button(button_frame, text="Confirmar", command=confirmar, style="TButton").pack(side="left", padx=10)
        ttk.Button(button_frame, text="Cancelar", command=w.destroy, style="TButton").pack(side="left", padx=10)

if __name__ == "__main__":
    import tkinter as tk
    import ctypes
    import traceback
    import tkinter.messagebox as mb

    def resource_path(relative_path):
        import os, sys
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")

        return os.path.join(base_path, relative_path)

    try:
        root = tk.Tk()
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(u"VillaVenecia.PayrollApp")
        root.iconbitmap(resource_path("nomina.ico"))
        app = PayrollApp(root)
        root.mainloop()

    except Exception as e:
        with open("error.log", "w", encoding="utf-8") as f:
            f.write(traceback.format_exc())
        mb.showerror("Error crítico", f"Ocurrió un error:\n{e}")



